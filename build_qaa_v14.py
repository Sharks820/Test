#!/usr/bin/env python3
"""
QAA Quote System v14.0 - Streamlined Edition
- Order # and PO # fields (optional)
- Fixed button alignment (no overlapping)
- No logo
- PO Tracking with YTD analysis
- Version tracking for inventory
- Infor XA safe import (no IT security triggers)
- Quick Reference Guide for each section
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from datetime import datetime

# Colors
NAVY = "1B365D"
SKY_BLUE = "00A3E0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
MEDIUM_GRAY = "E0E0E0"
DARK_GRAY = "666666"
GREEN = "28A745"
RED = "DC3545"
INPUT_YELLOW = "FFFDE7"

# Borders
thin_border = Border(
    left=Side(style='thin', color=MEDIUM_GRAY),
    right=Side(style='thin', color=MEDIUM_GRAY),
    top=Side(style='thin', color=MEDIUM_GRAY),
    bottom=Side(style='thin', color=MEDIUM_GRAY)
)

def apply_header(cell):
    cell.font = Font(bold=True, size=10, color=WHITE)
    cell.fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

def apply_input(cell):
    cell.fill = PatternFill(start_color=INPUT_YELLOW, end_color=INPUT_YELLOW, fill_type='solid')
    cell.border = thin_border

def apply_data(cell):
    cell.border = thin_border

# =============================================================================
# MENU SHEET
# =============================================================================
def create_menu_sheet(wb):
    ws = wb.create_sheet("Menu", 0)

    for col, w in {'A': 5, 'B': 20, 'C': 20, 'D': 20, 'E': 20, 'F': 5}.items():
        ws.column_dimensions[col].width = w

    # Header
    ws.merge_cells('B2:E2')
    ws['B2'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B2'].font = Font(size=20, bold=True, color=NAVY)
    ws['B2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B3:E3')
    ws['B3'] = "A Hartzell Engine Technologies Company"
    ws['B3'].font = Font(size=10, italic=True, color=SKY_BLUE)
    ws['B3'].alignment = Alignment(horizontal='center')

    # System title
    ws.merge_cells('B5:E5')
    ws['B5'] = "QUOTE SYSTEM v14.0"
    ws['B5'].font = Font(size=16, bold=True, color=WHITE)
    ws['B5'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 35

    # Quote types
    ws['B7'] = "QUOTE TYPES:"
    ws['B7'].font = Font(bold=True, color=NAVY)

    for col, (title, desc) in [
        ('B', ('500HR SERVICE', 'Magneto inspections')),
        ('C', ('CUST PROPERTY', 'Customer overhauls')),
        ('D', ('EXCHANGE/SALE', 'Exchange & inventory'))
    ]:
        ws[f'{col}8'] = title
        ws[f'{col}8'].font = Font(bold=True, size=11, color=NAVY)
        ws[f'{col}8'].alignment = Alignment(horizontal='center')
        ws[f'{col}9'] = desc
        ws[f'{col}9'].font = Font(size=9, color=DARK_GRAY)
        ws[f'{col}9'].alignment = Alignment(horizontal='center')

    # Button row - CLEARLY LABELED (buttons assigned in VBA)
    ws['B10'] = "[ Go500HR ]"
    ws['C10'] = "[ GoCustProp ]"
    ws['D10'] = "[ GoExchange ]"
    for col in ['B', 'C', 'D']:
        ws[f'{col}10'].font = Font(size=10, color=NAVY)
        ws[f'{col}10'].alignment = Alignment(horizontal='center')
        ws[f'{col}10'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
        ws[f'{col}10'].border = thin_border

    # Tools section
    ws.merge_cells('B12:E12')
    ws['B12'] = "TOOLS & TRACKING"
    ws['B12'].font = Font(size=14, bold=True, color=WHITE)
    ws['B12'].fill = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type='solid')
    ws['B12'].alignment = Alignment(horizontal='center')

    tools = [
        ('B13', 'Quote Tracker', 'Track follow-ups & status'),
        ('C13', 'PO Tracking', 'Order dates & YTD'),
        ('D13', 'Infor XA Data', 'Import inventory'),
        ('E13', 'Quick Reference', 'How-to guides'),
    ]
    for cell, title, desc in tools:
        ws[cell] = title
        ws[cell].font = Font(bold=True, color=NAVY)
        ws[cell].alignment = Alignment(horizontal='center')
        row = int(cell[1:]) + 1
        col = cell[0]
        ws[f'{col}{row}'] = desc
        ws[f'{col}{row}'].font = Font(size=9, color=DARK_GRAY)
        ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

    # Footer
    ws.merge_cells('B17:E17')
    ws['B17'] = "5746 E Apache St, Tulsa, OK 74115 | 918-835-6948 | sales@qaainc.com"
    ws['B17'].font = Font(size=9, color=DARK_GRAY)
    ws['B17'].alignment = Alignment(horizontal='center')

    return ws

# =============================================================================
# 500HR SHEET - With Order # and PO #
# =============================================================================
def create_500hr_sheet(wb):
    ws = wb.create_sheet("500HR")

    widths = {'A': 3, 'B': 15, 'C': 12, 'D': 26, 'E': 11, 'F': 9, 'G': 13, 'H': 11, 'I': 11, 'J': 3}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header
    ws.merge_cells('B1:I1')
    ws['B1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B1'].font = Font(size=16, bold=True, color=NAVY)
    ws['B1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B2:I2')
    ws['B2'] = "500HR MAGNETO INSPECTION SERVICE QUOTE"
    apply_header(ws['B2'])
    ws['B2'].font = Font(size=12, bold=True, color=WHITE)
    ws.row_dimensions[2].height = 28

    # Quote info row 1
    ws['B4'] = "Quote #:"
    ws['B4'].font = Font(bold=True, size=10, color=NAVY)
    ws['C4'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['C4'].font = Font(bold=True, size=11, color=NAVY)

    ws['E4'] = "Date:"
    ws['E4'].font = Font(bold=True, color=NAVY)
    ws['F4'] = datetime.now().strftime("%m/%d/%Y")

    ws['G4'] = "Valid:"
    ws['G4'].font = Font(bold=True, color=NAVY)
    ws['H4'] = "30 Days"

    # Quote info row 2 - ORDER # AND PO # (NEW!)
    ws['B5'] = "Order #:"
    ws['B5'].font = Font(bold=True, color=NAVY)
    ws['C5'] = ""
    apply_input(ws['C5'])

    ws['E5'] = "Customer PO #:"
    ws['E5'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('F5:G5')
    ws['F5'] = ""
    apply_input(ws['F5'])

    # Customer info row
    ws['B6'] = "Customer Email:"
    ws['B6'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('C6:E6')
    ws['C6'] = ""
    apply_input(ws['C6'])

    ws['F6'] = "Approval?"
    ws['F6'].font = Font(bold=True, size=9, color=NAVY)
    ws['G6'] = "No"
    apply_input(ws['G6'])

    dv_yesno = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(dv_yesno)
    dv_yesno.add(ws['G6'])

    # Separator
    ws.row_dimensions[7].height = 6
    for col in range(2, 10):
        ws.cell(row=7, column=col).border = Border(bottom=Side(style='thin', color=NAVY))

    # Magneto table header
    headers = ['Magneto P/N', 'Serial #', 'Description', '500HR $', 'Ovhl?', 'CO Part#', 'CO $', 'Total']
    for i, h in enumerate(headers):
        cell = ws.cell(row=8, column=i+2, value=h)
        apply_header(cell)
    ws.row_dimensions[8].height = 24

    # Magneto rows (6 units)
    for row in range(9, 15):
        # P/N
        ws[f'B{row}'] = ""
        apply_input(ws[f'B{row}'])

        # Serial
        ws[f'C{row}'] = ""
        apply_input(ws[f'C{row}'])

        # Description - 5H SUFFIX (NOT -5H!)
        ws[f'D{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row}&"5H",\'Parts Master\'!$A:$H,2,FALSE),IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found")))'
        apply_data(ws[f'D{row}'])

        # 500HR Price - 5H SUFFIX
        ws[f'E{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row}&"5H",\'Parts Master\'!$A:$H,4,FALSE),Config!$B$5))'
        apply_data(ws[f'E{row}'])
        ws[f'E{row}'].number_format = '$#,##0.00'

        # Overhaul?
        ws[f'F{row}'] = ""
        apply_input(ws[f'F{row}'])
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        dv_yesno.add(ws[f'F{row}'])

        # CO Part#
        ws[f'G{row}'] = f'=IF(F{row}="Yes",SUBSTITUTE(B{row},"5H","")&"-CO","")'
        apply_data(ws[f'G{row}'])

        # CO $
        ws[f'H{row}'] = f'=IF(G{row}="","",IFERROR(VLOOKUP(G{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        apply_data(ws[f'H{row}'])
        ws[f'H{row}'].number_format = '$#,##0.00'

        # Line Total
        ws[f'I{row}'] = f'=IF(B{row}="","",IF(F{row}="Yes",H{row},E{row}))'
        apply_data(ws[f'I{row}'])
        ws[f'I{row}'].number_format = '$#,##0.00'

    # Findings section
    ws.merge_cells('B16:I16')
    ws['B16'] = "INSPECTION FINDINGS"
    ws['B16'].font = Font(bold=True, size=10, color=WHITE)
    ws['B16'].fill = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type='solid')
    ws['B16'].alignment = Alignment(horizontal='center')

    findings = ["Coil", "Cap", "Dist", "Impulse", "Gear", "Points", "Brg", "Oil"]
    for i, f in enumerate(findings):
        cell = ws.cell(row=17, column=i+2, value=f)
        cell.font = Font(size=8, bold=True)
        cell.alignment = Alignment(horizontal='center')

    for row in range(18, 24):
        for col in range(2, 10):
            cell = ws.cell(row=row, column=col, value="☐")
            cell.font = Font(size=12)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Totals
    ws['B25'] = "Findings Fee:"
    ws['D25'] = '=SUMPRODUCT((B18:I23<>"☐")*1)/8*Config!$B$6'
    ws['D25'].number_format = '$#,##0.00'

    ws['F26'] = "Service Total:"
    ws['F26'].font = Font(bold=True, color=NAVY)
    ws['I26'] = '=SUMIF(F9:F14,"<>Yes",E9:E14)'
    ws['I26'].number_format = '$#,##0.00'

    ws['F27'] = "Overhaul Total:"
    ws['F27'].font = Font(bold=True, color=NAVY)
    ws['I27'] = '=SUMIF(F9:F14,"Yes",H9:H14)'
    ws['I27'].number_format = '$#,##0.00'

    ws.merge_cells('F28:G28')
    ws['F28'] = "GRAND TOTAL:"
    ws['F28'].font = Font(bold=True, size=12, color=NAVY)
    ws['F28'].alignment = Alignment(horizontal='right')

    ws.merge_cells('H28:I28')
    ws['H28'] = '=I26+I27+D25'
    ws['H28'].font = Font(bold=True, size=14, color=NAVY)
    ws['H28'].fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    ws['H28'].number_format = '$#,##0.00'
    ws['H28'].border = Border(
        left=Side(style='medium', color=NAVY),
        right=Side(style='medium', color=NAVY),
        top=Side(style='medium', color=NAVY),
        bottom=Side(style='medium', color=NAVY)
    )

    # Notes
    ws['B30'] = "Notes:"
    ws['B30'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('B31:I32')
    ws['B31'] = ""
    apply_input(ws['B31'])

    # Footer
    ws.merge_cells('B34:I34')
    ws['B34'] = "5746 E Apache St, Tulsa, OK 74115 | 918-835-6948 | sales@qaainc.com | qaa.com"
    ws['B34'].font = Font(size=9, color=DARK_GRAY)
    ws['B34'].alignment = Alignment(horizontal='center')

    # BUTTON ROW - PROPERLY SPACED (Row 36)
    ws.row_dimensions[36].height = 30
    buttons = [
        ('B36', '[ GENERATE PDF ]', 'GenPDF500HR'),
        ('D36', '[ SEND EMAIL ]', 'Email500HR'),
        ('F36', '[ CLEAR FORM ]', 'Clear500HR'),
        ('H36', '[ ← MENU ]', 'GoMenu'),
    ]
    for cell, label, macro in buttons:
        ws[cell] = label
        ws[cell].font = Font(bold=True, size=10, color=WHITE)
        ws[cell].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
        ws[cell].border = thin_border

    ws.print_area = 'A1:J34'

    return ws

# =============================================================================
# CUSTOMER PROPERTY SHEET - With Order # and PO #
# =============================================================================
def create_cust_property_sheet(wb):
    ws = wb.create_sheet("Cust Property")

    widths = {'A': 6, 'B': 15, 'C': 22, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 8, 'I': 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header
    ws.merge_cells('B1:H1')
    ws['B1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B1'].font = Font(size=16, bold=True, color=NAVY)
    ws['B1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B2:H2')
    ws['B2'] = "CUSTOMER PROPERTY OVERHAUL QUOTE"
    apply_header(ws['B2'])
    ws['B2'].font = Font(size=12, bold=True, color=WHITE)
    ws.row_dimensions[2].height = 28

    # Quote info row 1
    ws['B4'] = "Quote #:"
    ws['B4'].font = Font(bold=True, color=NAVY)
    ws['C4'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['C4'].font = Font(bold=True, size=11, color=NAVY)

    ws['E4'] = "Date:"
    ws['E4'].font = Font(bold=True, color=NAVY)
    ws['F4'] = datetime.now().strftime("%m/%d/%Y")

    ws['G4'] = "Updated:"
    ws['G4'].font = Font(bold=True, color=NAVY)
    ws['H4'] = ""
    apply_input(ws['H4'])

    # Quote info row 2 - ORDER # AND PO # (NEW!)
    ws['B5'] = "Order #:"
    ws['B5'].font = Font(bold=True, color=NAVY)
    ws['C5'] = ""
    apply_input(ws['C5'])

    ws['E5'] = "Customer PO #:"
    ws['E5'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('F5:G5')
    ws['F5'] = ""
    apply_input(ws['F5'])

    # Customer info
    ws['B6'] = "Customer Email:"
    ws['B6'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('C6:E6')
    ws['C6'] = ""
    apply_input(ws['C6'])

    ws['F6'] = "Tier:"
    ws['F6'].font = Font(bold=True, color=NAVY)
    ws['G6'] = "Standard"
    apply_input(ws['G6'])

    dv_tier = DataValidation(type="list", formula1='"Standard,Tier 1,Tier 2,Tier 3"')
    ws.add_data_validation(dv_tier)
    dv_tier.add(ws['G6'])

    # Needs row
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(dv_yesno)

    ws['B7'] = "Needs:"
    ws['B7'].font = Font(bold=True, size=9, color=NAVY)

    ws['C7'] = "Approval?"
    ws['D7'] = "No"
    apply_input(ws['D7'])
    dv_yesno.add(ws['D7'])

    ws['E7'] = "Ship?"
    ws['F7'] = "No"
    apply_input(ws['F7'])
    dv_yesno.add(ws['F7'])

    ws['G7'] = "Pay?"
    ws['H7'] = "No"
    apply_input(ws['H7'])
    dv_yesno.add(ws['H7'])

    # Separator
    ws.row_dimensions[8].height = 6

    # Units table
    headers = ['Unit P/N', 'Description', 'Type', 'Base $', 'Your $', 'Exch Cap']
    for i, h in enumerate(headers):
        cell = ws.cell(row=9, column=i+2, value=h)
        apply_header(cell)

    for row in range(10, 14):
        ws[f'B{row}'] = ""
        apply_input(ws[f'B{row}'])

        ws[f'C{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found"))'
        apply_data(ws[f'C{row}'])

        ws[f'D{row}'] = f'=IF(B{row}="","",IF(COUNTIF(BOM!$A:$A,SUBSTITUTE(B{row},"-CO",""))>0,"In-House","Vendor"))'
        apply_data(ws[f'D{row}'])
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        ws[f'E{row}'] = ""
        apply_input(ws[f'E{row}'])
        ws[f'E{row}'].number_format = '$#,##0.00'

        ws[f'F{row}'] = f'=IF(E{row}="","",E{row}*(1-VLOOKUP($G$6,Config!$A$10:$B$13,2,FALSE)))'
        apply_data(ws[f'F{row}'])
        ws[f'F{row}'].number_format = '$#,##0.00'

        ws[f'G{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(SUBSTITUTE(B{row},"-CO","-OH"),\'Parts Master\'!$A:$H,4,FALSE),"N/A"))'
        apply_data(ws[f'G{row}'])
        ws[f'G{row}'].number_format = '$#,##0.00'

    # O&A section
    ws.merge_cells('B15:H15')
    ws['B15'] = "OVER & ABOVE PARTS"
    ws['B15'].font = Font(bold=True, color=WHITE)
    ws['B15'].fill = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type='solid')
    ws['B15'].alignment = Alignment(horizontal='center')

    oa_headers = ['#', 'Part #', 'Description', 'Qty', 'Unit $', 'Ext $', 'Status']
    for i, h in enumerate(oa_headers):
        cell = ws.cell(row=16, column=i+1, value=h)
        cell.font = Font(bold=True, size=9, color=NAVY)
        cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    for row in range(17, 25):
        ws[f'A{row}'] = ""
        apply_input(ws[f'A{row}'])

        ws[f'B{row}'] = ""
        apply_input(ws[f'B{row}'])

        ws[f'C{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),""))'
        apply_data(ws[f'C{row}'])

        ws[f'D{row}'] = ""
        apply_input(ws[f'D{row}'])
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        ws[f'E{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        apply_data(ws[f'E{row}'])
        ws[f'E{row}'].number_format = '$#,##0.00'

        ws[f'F{row}'] = f'=IF(D{row}="","",E{row}*D{row})'
        apply_data(ws[f'F{row}'])
        ws[f'F{row}'].number_format = '$#,##0.00'

        ws[f'G{row}'] = ""
        apply_input(ws[f'G{row}'])

    # Totals
    ws['E26'] = "Labor:"
    ws['E26'].font = Font(bold=True, color=NAVY)
    ws['F26'] = '=SUM(F10:F13)'
    ws['F26'].number_format = '$#,##0.00'

    ws['E27'] = "Parts:"
    ws['E27'].font = Font(bold=True, color=NAVY)
    ws['F27'] = '=SUM(F17:F24)'
    ws['F27'].number_format = '$#,##0.00'

    ws['E28'] = "TOTAL:"
    ws['E28'].font = Font(bold=True, size=12, color=NAVY)
    ws['F28'] = '=F26+F27'
    ws['F28'].font = Font(bold=True, size=12, color=NAVY)
    ws['F28'].fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    ws['F28'].number_format = '$#,##0.00'

    # Notes
    ws['B30'] = "Notes:"
    ws.merge_cells('B31:G32')
    ws['B31'] = ""
    apply_input(ws['B31'])

    # Footer
    ws.merge_cells('B34:G34')
    ws['B34'] = "918-835-6948 | sales@qaainc.com | qaa.com"
    ws['B34'].font = Font(size=9, color=DARK_GRAY)
    ws['B34'].alignment = Alignment(horizontal='center')

    # BUTTON ROW - PROPERLY SPACED
    ws.row_dimensions[36].height = 30
    for cell, label in [('B36', '[ GENERATE PDF ]'), ('D36', '[ SEND EMAIL ]'),
                         ('F36', '[ CLEAR ]'), ('H36', '[ MENU ]')]:
        ws[cell] = label
        ws[cell].font = Font(bold=True, size=9, color=WHITE)
        ws[cell].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    ws.print_area = 'A1:H34'

    return ws

# =============================================================================
# EXCHANGE SHEET
# =============================================================================
def create_exchange_sheet(wb):
    ws = wb.create_sheet("Exchange")

    widths = {'A': 15, 'B': 22, 'C': 6, 'D': 5, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 7, 'J': 9}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['A1'].font = Font(size=16, bold=True, color=NAVY)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = "EXCHANGE / INVENTORY SALE QUOTE"
    apply_header(ws['A2'])
    ws['A2'].font = Font(size=12, bold=True, color=WHITE)
    ws.row_dimensions[2].height = 28

    # Quote info
    ws['A4'] = "Quote #:"
    ws['A4'].font = Font(bold=True, color=NAVY)
    ws['B4'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['B4'].font = Font(bold=True, color=NAVY)

    ws['D4'] = "Date:"
    ws['E4'] = datetime.now().strftime("%m/%d/%Y")

    ws['G4'] = "Valid: 30 Days"

    # Customer
    ws['A5'] = "Customer Email:"
    ws['A5'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('B5:D5')
    ws['B5'] = ""
    apply_input(ws['B5'])

    ws['F5'] = "Tier:"
    ws['G5'] = "Standard"
    apply_input(ws['G5'])

    dv_tier = DataValidation(type="list", formula1='"Standard,Tier 1,Tier 2,Tier 3"')
    ws.add_data_validation(dv_tier)
    dv_tier.add(ws['G5'])

    # Requirements
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(dv_yesno)

    ws['A6'] = "Approval?"
    ws['B6'] = "No"
    apply_input(ws['B6'])
    dv_yesno.add(ws['B6'])

    ws['C6'] = "Ship?"
    ws['D6'] = "No"
    apply_input(ws['D6'])
    dv_yesno.add(ws['D6'])

    ws['E6'] = "Pay?"
    ws['F6'] = "No"
    apply_input(ws['F6'])
    dv_yesno.add(ws['F6'])

    ws['G6'] = "Core In Hand?"
    ws['H6'] = "No"
    apply_input(ws['H6'])
    dv_yesno.add(ws['H6'])

    # Line items header
    headers = ['Part #', 'Description', 'Cond', 'Qty', 'List $', 'Your $', 'Ext $', 'Core $', 'Stock', 'Lead']
    for i, h in enumerate(headers):
        cell = ws.cell(row=8, column=i+1, value=h)
        apply_header(cell)

    # Line items (10 rows)
    for row in range(9, 19):
        ws[f'A{row}'] = ""
        apply_input(ws[f'A{row}'])

        ws[f'B{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found"))'
        apply_data(ws[f'B{row}'])

        ws[f'C{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,3,FALSE),""))'
        apply_data(ws[f'C{row}'])

        ws[f'D{row}'] = ""
        apply_input(ws[f'D{row}'])
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        ws[f'E{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        apply_data(ws[f'E{row}'])
        ws[f'E{row}'].number_format = '$#,##0.00'

        ws[f'F{row}'] = f'=IF(E{row}="","",E{row}*(1-VLOOKUP($G$5,Config!$A$10:$B$13,2,FALSE)))'
        apply_data(ws[f'F{row}'])
        ws[f'F{row}'].number_format = '$#,##0.00'

        ws[f'G{row}'] = f'=IF(D{row}="","",F{row}*D{row})'
        apply_data(ws[f'G{row}'])
        ws[f'G{row}'].number_format = '$#,##0.00'

        ws[f'H{row}'] = f'=IF(A{row}="","",IF($H$6="Yes",0,IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,5,FALSE),0)*D{row}))'
        apply_data(ws[f'H{row}'])
        ws[f'H{row}'].number_format = '$#,##0.00'

        ws[f'I{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,6,FALSE),""))'
        apply_data(ws[f'I{row}'])

        ws[f'J{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,7,FALSE),""))'
        apply_data(ws[f'J{row}'])

    # Totals
    ws['A20'] = "Subtotal:"
    ws['B20'] = '=SUM(G9:G18)'
    ws['B20'].number_format = '$#,##0.00'
    ws['B20'].font = Font(bold=True)

    ws['D20'] = "Core:"
    ws['E20'] = '=SUM(H9:H18)'
    ws['E20'].number_format = '$#,##0.00'

    ws['F20'] = '=IF(H6="Yes","(WAIVED)","")'
    ws['F20'].font = Font(color=GREEN)

    ws['G20'] = "Savings:"
    ws['H20'] = '=SUM(E9:E18)-SUM(F9:F18)'
    ws['H20'].font = Font(bold=True, color=GREEN)
    ws['H20'].number_format = '$#,##0.00'

    ws.merge_cells('D21:E21')
    ws['D21'] = "GRAND TOTAL:"
    ws['D21'].font = Font(bold=True, size=12, color=NAVY)
    ws['D21'].alignment = Alignment(horizontal='right')

    ws.merge_cells('F21:G21')
    ws['F21'] = '=B20+IF(H6="Yes",0,E20)'
    ws['F21'].font = Font(bold=True, size=14, color=NAVY)
    ws['F21'].fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    ws['F21'].number_format = '$#,##0.00'

    # Notes
    ws['A23'] = "Notes:"
    ws.merge_cells('A24:J25')
    ws['A24'] = ""
    apply_input(ws['A24'])

    # Footer
    ws.merge_cells('A27:J27')
    ws['A27'] = "Core charges refunded upon receipt of serviceable core within 30 days."
    ws['A27'].font = Font(size=9, italic=True, color=DARK_GRAY)
    ws['A27'].alignment = Alignment(horizontal='center')

    # BUTTON ROW
    ws.row_dimensions[29].height = 30
    for cell, label in [('A29', '[ PDF ]'), ('C29', '[ EMAIL ]'), ('E29', '[ CLEAR ]'), ('G29', '[ MENU ]')]:
        ws[cell] = label
        ws[cell].font = Font(bold=True, size=9, color=WHITE)
        ws[cell].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')

    ws.print_area = 'A1:J27'
    ws.page_setup.orientation = 'landscape'

    return ws

# =============================================================================
# QUOTE TRACKER - Single Intelligent Log
# =============================================================================
def create_quote_tracker(wb):
    ws = wb.create_sheet("Quote Tracker")

    widths = {'A': 12, 'B': 10, 'C': 22, 'D': 11, 'E': 10, 'F': 9, 'G': 9, 'H': 9,
              'I': 8, 'J': 10, 'K': 8, 'L': 9, 'M': 35}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = "QUOTE TRACKER - Follow-up & Status Management"
    ws['A1'].font = Font(size=14, bold=True, color=WHITE)
    ws['A1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Instructions
    ws.merge_cells('A2:M2')
    ws['A2'] = "RED = needs follow-up (7+ days) | Mark needs as DONE when fulfilled | Set Status to CLOSED when complete"
    ws['A2'].font = Font(size=9, italic=True, color=DARK_GRAY)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = ['Quote #', 'Created', 'Customer Email', 'Type', 'Amount', 'Approval',
               'Shipping', 'Payment', 'Reminders', 'Last Contact', 'Days', 'Status', 'Notes']
    for i, h in enumerate(headers):
        cell = ws.cell(row=3, column=i+1, value=h)
        apply_header(cell)

    # Data validation
    dv_needs = DataValidation(type="list", formula1='"Yes,No,DONE"')
    dv_status = DataValidation(type="list", formula1='"OPEN,CLOSED,ON HOLD"')
    ws.add_data_validation(dv_needs)
    ws.add_data_validation(dv_status)

    # Data rows
    for row in range(4, 54):
        for col in range(1, 14):
            ws.cell(row=row, column=col).border = thin_border

        ws[f'B{row}'].number_format = 'MM/DD/YY'
        ws[f'E{row}'].number_format = '$#,##0.00'
        ws[f'J{row}'].number_format = 'MM/DD/YY'

        for col in ['D', 'F', 'G', 'H', 'I', 'K', 'L']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')

        for col in ['F', 'G', 'H']:
            dv_needs.add(ws[f'{col}{row}'])

        dv_status.add(ws[f'L{row}'])

        # Days Since formula
        ws[f'K{row}'] = f'=IF(J{row}="","",TODAY()-J{row})'

        ws[f'M{row}'].alignment = Alignment(wrap_text=True)

    # Conditional formatting
    red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

    ws.conditional_formatting.add('A4:M53', FormulaRule(formula=['AND($L4="OPEN",$K4>=7)'], fill=red_fill))
    ws.conditional_formatting.add('A4:M53', FormulaRule(formula=['AND($L4="OPEN",$K4>=3,$K4<7)'], fill=yellow_fill))
    ws.conditional_formatting.add('A4:M53', FormulaRule(formula=['$L4="CLOSED"'], fill=green_fill))

    # Summary
    ws['A55'] = "Open:"
    ws['B55'] = '=COUNTIF(L4:L53,"OPEN")'
    ws['C55'] = "Need Follow-up:"
    ws['D55'] = '=SUMPRODUCT((L4:L53="OPEN")*(K4:K53>=7))'
    ws['D55'].font = Font(bold=True, color=RED)

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = 'A3:M53'

    return ws

# =============================================================================
# PO TRACKING - Create Date to Complete Date YTD
# =============================================================================
def create_po_tracking(wb):
    ws = wb.create_sheet("PO Tracking")

    widths = {'A': 12, 'B': 14, 'C': 20, 'D': 12, 'E': 12, 'F': 12, 'G': 10, 'H': 25}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "PO TRACKING - Create Date to Complete Date Analysis"
    ws['A1'].font = Font(size=14, bold=True, color=WHITE)
    ws['A1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # YTD Summary section
    ws['A3'] = "YTD SUMMARY"
    ws['A3'].font = Font(bold=True, size=12, color=NAVY)

    ws['A4'] = "Current Year:"
    ws['B4'] = '=YEAR(TODAY())'
    ws['C4'] = "Previous Year:"
    ws['D4'] = '=YEAR(TODAY())-1'

    ws['A5'] = "YTD Orders:"
    ws['B5'] = '=COUNTIFS(D9:D500,">="&DATE(YEAR(TODAY()),1,1),D9:D500,"<="&TODAY())'
    ws['C5'] = "YTD Completed:"
    ws['D5'] = '=COUNTIFS(E9:E500,">="&DATE(YEAR(TODAY()),1,1),E9:E500,"<="&TODAY())'

    ws['A6'] = "Avg Days to Complete:"
    ws['B6'] = '=IFERROR(AVERAGEIFS(F9:F500,F9:F500,">0",D9:D500,">="&DATE(YEAR(TODAY()),1,1)),0)'
    ws['B6'].number_format = '0.0'

    # Headers
    headers = ['PO #', 'Order #', 'Customer', 'Create Date', 'Complete Date', 'Days', 'Status', 'Notes']
    for i, h in enumerate(headers):
        cell = ws.cell(row=8, column=i+1, value=h)
        apply_header(cell)

    # Data validation
    dv_status = DataValidation(type="list", formula1='"Open,In Process,Complete,Cancelled"')
    ws.add_data_validation(dv_status)

    # Data rows
    for row in range(9, 109):
        for col in range(1, 9):
            ws.cell(row=row, column=col).border = thin_border

        ws[f'D{row}'].number_format = 'MM/DD/YYYY'
        ws[f'E{row}'].number_format = 'MM/DD/YYYY'

        # Days calculation
        ws[f'F{row}'] = f'=IF(AND(D{row}<>"",E{row}<>""),E{row}-D{row},IF(D{row}<>"",TODAY()-D{row},""))'
        ws[f'F{row}'].alignment = Alignment(horizontal='center')

        dv_status.add(ws[f'G{row}'])
        ws[f'G{row}'].alignment = Alignment(horizontal='center')

    # Conditional formatting for overdue
    red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    ws.conditional_formatting.add('F9:F108', FormulaRule(formula=['AND($G9<>"Complete",$F9>14)'], fill=red_fill))

    ws.freeze_panes = 'A9'

    return ws

# =============================================================================
# INFOR XA DATA - Safe Import (No IT Security Triggers)
# =============================================================================
def create_infor_xa_sheet(wb):
    ws = wb.create_sheet("Infor XA Data")

    widths = {'A': 15, 'B': 30, 'C': 8, 'D': 10, 'E': 10, 'F': 10, 'G': 12, 'H': 12, 'I': 20}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "INFOR XA / MAPICS DATA IMPORT"
    ws['A1'].font = Font(size=14, bold=True, color=WHITE)
    ws['A1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Instructions
    ws.merge_cells('A3:I3')
    ws['A3'] = "HOW TO UPDATE: Copy data from Infor XA Power Inquiry → Paste below → Data refreshes Parts Master"
    ws['A3'].font = Font(size=10, bold=True, color=NAVY)

    ws.merge_cells('A4:I4')
    ws['A4'] = "No direct database connection = No IT security concerns. Manual paste is safe and controlled."
    ws['A4'].font = Font(size=9, italic=True, color=DARK_GRAY)

    # Version tracking
    ws['A6'] = "Last Updated:"
    ws['A6'].font = Font(bold=True, color=NAVY)
    ws['B6'] = ""
    apply_input(ws['B6'])
    ws['B6'].number_format = 'MM/DD/YYYY HH:MM'

    ws['C6'] = "Updated By:"
    ws['D6'] = ""
    apply_input(ws['D6'])

    ws['E6'] = "Version:"
    ws['F6'] = ""
    apply_input(ws['F6'])

    # Headers
    headers = ['Item #', 'Description', 'U/M', 'On Hand', 'Allocated', 'Available', 'Last Cost', 'List Price', 'Location']
    for i, h in enumerate(headers):
        cell = ws.cell(row=8, column=i+1, value=h)
        apply_header(cell)

    # Paste area (100 rows)
    for row in range(9, 109):
        for col in range(1, 10):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if col in [4, 5, 6]:
                cell.alignment = Alignment(horizontal='center')
            if col in [7, 8]:
                cell.number_format = '$#,##0.00'

    # Highlight paste area
    ws.merge_cells('A109:I109')
    ws['A109'] = "← Paste Infor XA data above (up to 100 rows) →"
    ws['A109'].font = Font(italic=True, color=DARK_GRAY)
    ws['A109'].alignment = Alignment(horizontal='center')

    return ws

# =============================================================================
# QUICK REFERENCE GUIDE
# =============================================================================
def create_quick_reference(wb):
    ws = wb.create_sheet("Quick Reference")

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 3

    # Title
    ws.merge_cells('B1:C1')
    ws['B1'] = "QUICK REFERENCE GUIDE"
    ws['B1'].font = Font(size=16, bold=True, color=WHITE)
    ws['B1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35

    row = 3

    sections = [
        ("500HR SERVICE", [
            ("Enter Magneto P/N", "Type part number in column B (e.g., 10-382320)"),
            ("Price Lookup", "System auto-looks up 5H variant for 500HR price"),
            ("Mark Overhaul", "Set 'Ovhl?' to Yes if unit needs complete overhaul"),
            ("Findings", "Click checkboxes ☐→☑ for any issues found"),
            ("Generate PDF", "Click button or run GenPDF500HR macro"),
            ("Send Email", "Click button - auto-logs to Quote Tracker"),
        ]),
        ("CUSTOMER PROPERTY", [
            ("Enter Unit P/N", "Use -CO suffix for customer overhaul parts"),
            ("Enter Base Price", "Manual entry - your labor quote"),
            ("Tier Discount", "Select customer tier for automatic discount"),
            ("O&A Parts", "Add over-and-above parts with quantities"),
            ("Order # / PO #", "Optional - shows on PDF if entered"),
        ]),
        ("EXCHANGE / SALE", [
            ("Enter Part #", "Use -OH suffix for overhauled exchange units"),
            ("Core In Hand", "Set to Yes to waive core charges"),
            ("Quantity", "Enter qty - extended price auto-calculates"),
            ("Stock/Lead", "Auto-pulled from Parts Master"),
        ]),
        ("QUOTE TRACKER", [
            ("Status", "OPEN = active, CLOSED = complete, ON HOLD = waiting"),
            ("Needs Columns", "Yes = needed, No = not needed, DONE = received"),
            ("Reminders", "Count of follow-up contacts made"),
            ("RED Row", "7+ days since contact - needs follow-up!"),
            ("YELLOW Row", "3-6 days - approaching follow-up time"),
            ("GREEN Row", "Quote is closed"),
        ]),
        ("PO TRACKING", [
            ("Create Date", "When PO was created/received"),
            ("Complete Date", "When order shipped/completed"),
            ("Days", "Auto-calculated turnaround time"),
            ("YTD Stats", "Summary shows current year performance"),
        ]),
        ("INFOR XA DATA", [
            ("Purpose", "Import inventory data without IT security issues"),
            ("How To", "Run Power Inquiry in XA → Copy results → Paste here"),
            ("Version", "Update 'Last Updated' and 'Version' after paste"),
            ("Refresh", "Parts Master can reference this for live stock levels"),
        ]),
    ]

    for section_title, items in sections:
        ws[f'B{row}'] = section_title
        ws[f'B{row}'].font = Font(bold=True, size=12, color=WHITE)
        ws[f'B{row}'].fill = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type='solid')
        ws.merge_cells(f'B{row}:C{row}')
        row += 1

        for action, description in items:
            ws[f'B{row}'] = action
            ws[f'B{row}'].font = Font(bold=True, color=NAVY)
            ws[f'C{row}'] = description
            ws[f'C{row}'].alignment = Alignment(wrap_text=True)
            row += 1

        row += 1  # Space between sections

    # Tips section
    ws[f'B{row}'] = "PRO TIPS"
    ws[f'B{row}'].font = Font(bold=True, size=12, color=WHITE)
    ws[f'B{row}'].fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
    ws.merge_cells(f'B{row}:C{row}')
    row += 1

    tips = [
        "Use Ctrl+; to enter today's date quickly",
        "Filter Quote Tracker by Status to see only OPEN quotes",
        "Check Quote Tracker weekly for RED rows needing follow-up",
        "Update Infor XA Data weekly for accurate inventory",
        "PO # and Order # are optional but helpful for tracking",
    ]

    for tip in tips:
        ws[f'B{row}'] = "•"
        ws[f'C{row}'] = tip
        row += 1

    return ws

# =============================================================================
# PARTS MASTER
# =============================================================================
def create_parts_master(wb):
    ws = wb.create_sheet("Parts Master")

    headers = ['Part Number', 'Description', 'Cond', 'List Price', 'Core Charge', 'Stock', 'Lead Time', 'Notes']
    widths = [16, 32, 8, 11, 11, 7, 11, 25]

    for i, (h, w) in enumerate(zip(headers, widths)):
        col = get_column_letter(i+1)
        ws[f'{col}1'] = h
        apply_header(ws[f'{col}1'])
        ws.column_dimensions[col].width = w

    # Data with 5H suffix (NOT -5H!)
    parts = [
        ['10-382320', 'S-20/200 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-3823205H', 'S-20/200 - 500HR Inspection', 'Svc', 575, 0, 0, '3-5 Days', '500HR'],
        ['10-382320-CO', 'S-20/200 - Cust Overhaul', 'CO', 895, 0, 0, '7-10 Days', ''],
        ['10-382320-OH', 'S-20/200 - OH Exchange', 'OH', 2495, 1000, 5, '5-7 Days', ''],

        ['10-320680', 'S-1200 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-3206805H', 'S-1200 - 500HR Inspection', 'Svc', 495, 0, 0, '3-5 Days', '500HR'],
        ['10-320680-CO', 'S-1200 - Cust Overhaul', 'CO', 750, 0, 0, '7-10 Days', ''],
        ['10-320680-OH', 'S-1200 - OH Exchange', 'OH', 2095, 800, 3, '5-7 Days', ''],

        ['10-391088', 'D-2000/D-3000 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-3910885H', 'D-2000/D-3000 - 500HR Inspection', 'Svc', 695, 0, 0, '3-5 Days', '500HR'],
        ['10-391088-CO', 'D-2000/D-3000 - Cust Overhaul', 'CO', 1150, 0, 0, '7-10 Days', ''],
        ['10-391088-OH', 'D-2000/D-3000 - OH Exchange', 'OH', 2850, 1100, 2, '5-7 Days', ''],

        ['10-163020', 'Bendix S-20 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-1630205H', 'Bendix S-20 - 500HR Inspection', 'Svc', 625, 0, 0, '3-5 Days', '500HR'],
        ['10-163020-CO', 'Bendix S-20 - Cust Overhaul', 'CO', 950, 0, 0, '7-10 Days', ''],
        ['10-163020-OH', 'Bendix S-20 - OH Exchange', 'OH', 2650, 950, 4, '5-7 Days', ''],

        ['LW-15473', 'Impulse Coupling Assembly', 'New', 425, 0, 12, 'In Stock', ''],
        ['10-357170', 'Coil Assembly', 'New', 185, 0, 20, 'In Stock', ''],
        ['10-357210', 'Capacitor', 'New', 45, 0, 50, 'In Stock', ''],
        ['10-51360', 'Distributor Block', 'New', 125, 0, 15, 'In Stock', ''],
        ['10-51350', 'Points Assembly', 'New', 85, 0, 30, 'In Stock', ''],
        ['10-51340', 'Bearing Set', 'New', 95, 0, 25, 'In Stock', ''],

        ['646942-CO', 'Governor - Cust Overhaul', 'CO', 1500, 0, 0, '14-21 Days', 'Vendor'],
        ['646942-OH', 'Governor - OH Exchange', 'OH', 3450, 1500, 2, '5-7 Days', ''],
        ['478615-CO', 'Turbocharger - Cust Overhaul', 'CO', 2800, 0, 0, '21-30 Days', 'Vendor'],
        ['478615-OH', 'Turbocharger - OH Exchange', 'OH', 5800, 2500, 1, '5-7 Days', ''],
    ]

    for r, row_data in enumerate(parts, start=2):
        for c, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in [4, 5]:
                cell.number_format = '$#,##0.00'

    return ws

# =============================================================================
# BOM SHEET
# =============================================================================
def create_bom(wb):
    ws = wb.create_sheet("BOM")

    headers = ['Parent P/N', 'Description', 'Component P/N', 'Comp Desc', 'Qty']
    for i, h in enumerate(headers):
        col = get_column_letter(i+1)
        ws[f'{col}1'] = h
        apply_header(ws[f'{col}1'])
        ws.column_dimensions[col].width = 22

    data = [
        ['10-382320', 'S-20/200 Magneto', '10-357170', 'Coil Assembly', 1],
        ['10-382320', 'S-20/200 Magneto', '10-357210', 'Capacitor', 1],
        ['10-382320', 'S-20/200 Magneto', '10-51360', 'Distributor Block', 1],
        ['10-320680', 'S-1200 Magneto', '10-357170', 'Coil Assembly', 1],
        ['10-391088', 'D-2000/D-3000', '10-357170', 'Coil Assembly', 1],
    ]

    for r, row_data in enumerate(data, start=2):
        for c, val in enumerate(row_data, start=1):
            ws.cell(row=r, column=c, value=val)

    return ws

# =============================================================================
# CONFIG SHEET
# =============================================================================
def create_config(wb):
    ws = wb.create_sheet("Config")

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30

    ws['A1'] = "SYSTEM CONFIGURATION"
    ws['A1'].font = Font(bold=True, size=14, color=NAVY)

    settings = [
        ('Quote Prefix', 'QAA', ''),
        ('Next Quote #', 1, 'Auto-increments'),
        ('Quote Valid Days', 30, ''),
        ('500HR Default Fee', 550, 'Fallback'),
        ('500HR Findings Fee', 110, 'Per unit'),
        ('Reminder Days', 7, 'Follow-up threshold'),
    ]

    for i, (name, val, note) in enumerate(settings, start=2):
        ws[f'A{i}'] = name
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = val
        apply_input(ws[f'B{i}'])
        ws[f'C{i}'] = note
        ws[f'C{i}'].font = Font(size=9, color=DARK_GRAY)

    ws['A9'] = "TIER DISCOUNTS"
    ws['A9'].font = Font(bold=True, size=12, color=NAVY)

    for i, (tier, disc) in enumerate([('Standard', 0), ('Tier 1', 0.07), ('Tier 2', 0.10), ('Tier 3', 0.13)], start=10):
        ws[f'A{i}'] = tier
        ws[f'B{i}'] = disc
        ws[f'B{i}'].number_format = '0%'

    # Version info
    ws['A15'] = "VERSION INFO"
    ws['A15'].font = Font(bold=True, size=12, color=NAVY)
    ws['A16'] = "System Version:"
    ws['B16'] = "14.0"
    ws['A17'] = "Last Updated:"
    ws['B17'] = datetime.now().strftime("%m/%d/%Y")

    return ws

# =============================================================================
# MAIN BUILD
# =============================================================================
def build_v14():
    print("=" * 60)
    print("QAA QUOTE SYSTEM v14.0 - Streamlined Edition")
    print("=" * 60)

    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print("\n[1/10] Creating Menu...")
    create_menu_sheet(wb)

    print("[2/10] Creating 500HR (with Order#/PO#)...")
    create_500hr_sheet(wb)

    print("[3/10] Creating Cust Property (with Order#/PO#)...")
    create_cust_property_sheet(wb)

    print("[4/10] Creating Exchange...")
    create_exchange_sheet(wb)

    print("[5/10] Creating Quote Tracker...")
    create_quote_tracker(wb)

    print("[6/10] Creating PO Tracking (YTD analysis)...")
    create_po_tracking(wb)

    print("[7/10] Creating Infor XA Data import...")
    create_infor_xa_sheet(wb)

    print("[8/10] Creating Quick Reference Guide...")
    create_quick_reference(wb)

    print("[9/10] Creating Parts Master & Config...")
    create_parts_master(wb)
    create_bom(wb)
    create_config(wb)

    print("[10/10] Saving...")
    output_path = "/home/user/Test/QAA_Quote_System_v14.xlsx"
    wb.save(output_path)
    print(f"\n  Saved: {output_path}")

    print("\n" + "=" * 60)
    print("BUILD COMPLETE!")
    print("=" * 60)
    print("\nNEW IN v14:")
    print("  - Order # and PO # fields (optional, show on PDF)")
    print("  - Fixed button alignment (no overlapping)")
    print("  - No logo (removed)")
    print("  - PO Tracking with YTD analysis")
    print("  - Infor XA safe data import")
    print("  - Quick Reference Guide")
    print("  - Version tracking in Config")

    return output_path

if __name__ == "__main__":
    build_v14()
