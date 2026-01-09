#!/usr/bin/env python3
"""
QAA Quote System Builder
Creates a complete Excel workbook with all sheets, formulas, images, and VBA code
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill,
                              NamedStyle, Protection)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.chart import BarChart, Reference
from PIL import Image, ImageDraw, ImageFont
import zipfile
import shutil
from datetime import datetime

# =============================================================================
# COLORS
# =============================================================================
NAVY = "1B365D"
SKY_BLUE = "00A3E0"
GOLD = "FFD966"
GREEN = "C6EFCE"
LIGHT_GRAY = "F2F2F2"
WHITE = "FFFFFF"
BLACK = "000000"

# =============================================================================
# STYLES
# =============================================================================
header_font = Font(name='Calibri', size=14, bold=True, color=WHITE)
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
subheader_font = Font(name='Calibri', size=11, bold=True, color=NAVY)
normal_font = Font(name='Calibri', size=10)
input_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type='solid')
total_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
medium_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

def create_button_images(output_dir):
    """Generate professional button PNG images"""

    buttons = {
        'btn_generate_pdf': ('GENERATE PDF', (180, 45)),
        'btn_email': ('EMAIL QUOTE', (150, 45)),
        'btn_clear': ('CLEAR FORM', (140, 45)),
        'btn_menu': ('◀ MENU', (100, 40)),
        'btn_update': ('UPDATE DATE', (140, 40)),
        'btn_500hr': ('500HR SERVICE', (160, 50)),
        'btn_custprop': ('CUSTOMER PROPERTY', (200, 50)),
        'btn_exchange': ('EXCHANGE / SALE', (180, 50)),
    }

    for name, (text, size) in buttons.items():
        img = Image.new('RGBA', size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # Navy blue rounded rectangle
        navy_rgb = (27, 54, 93)
        radius = 8

        # Draw rounded rectangle
        draw.rounded_rectangle(
            [(0, 0), (size[0]-1, size[1]-1)],
            radius=radius,
            fill=navy_rgb,
            outline=(0, 163, 224),
            width=2
        )

        # Add text
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
        except:
            font = ImageFont.load_default()

        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        x = (size[0] - text_width) // 2
        y = (size[1] - text_height) // 2 - 2

        draw.text((x, y), text, fill=(255, 255, 255), font=font)

        filepath = os.path.join(output_dir, f'{name}.png')
        img.save(filepath, 'PNG')
        print(f"Created: {filepath}")

    return buttons.keys()

def create_menu_sheet(wb, logo_path):
    """Create the Menu navigation sheet"""
    ws = wb.create_sheet("Menu", 0)

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 5

    # Header
    ws.merge_cells('B2:E2')
    ws['B2'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B2'].font = Font(name='Calibri', size=24, bold=True, color=NAVY)
    ws['B2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B3:E3')
    ws['B3'] = "A Hartzell Engine Technologies Company"
    ws['B3'].font = Font(name='Calibri', size=12, italic=True, color=SKY_BLUE)
    ws['B3'].alignment = Alignment(horizontal='center')

    # Logo placeholder
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 200
            img.height = 80
            ws.add_image(img, 'B5')
        except Exception as e:
            print(f"Logo error: {e}")

    # Title
    ws.merge_cells('B8:E8')
    ws['B8'] = "QUOTE SYSTEM"
    ws['B8'].font = Font(name='Calibri', size=20, bold=True, color=WHITE)
    ws['B8'].fill = header_fill
    ws['B8'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[8].height = 40

    # Navigation section labels
    ws['B11'] = "SELECT QUOTE TYPE:"
    ws['B11'].font = Font(name='Calibri', size=14, bold=True, color=NAVY)

    # Quote type descriptions
    ws['B13'] = "500HR SERVICE"
    ws['B13'].font = subheader_font
    ws['B14'] = "500-hour magneto inspection quotes"
    ws['B14'].font = Font(name='Calibri', size=10, color='666666')

    ws['C13'] = "CUSTOMER PROPERTY"
    ws['C13'].font = subheader_font
    ws['C14'] = "Overhaul quotes for customer units"
    ws['C14'].font = Font(name='Calibri', size=10, color='666666')

    ws['D13'] = "EXCHANGE / SALE"
    ws['D13'].font = subheader_font
    ws['D14'] = "Exchange units & inventory sales"
    ws['D14'].font = Font(name='Calibri', size=10, color='666666')

    # Footer
    ws.merge_cells('B20:E20')
    ws['B20'] = "5746 E Apache St, Tulsa, OK 74115 | 918-835-6948 | sales@qaainc.com | qaa.com"
    ws['B20'].font = Font(name='Calibri', size=10, color='666666')
    ws['B20'].alignment = Alignment(horizontal='center')

    # Button placement notes (will be replaced with actual buttons)
    ws['B16'] = "[500HR Button]"
    ws['C16'] = "[Cust Prop Button]"
    ws['D16'] = "[Exchange Button]"

    return ws

def create_500hr_sheet(wb):
    """Create the 500HR Magneto Service sheet with CORRECT 5H suffix formulas"""
    ws = wb.create_sheet("500HR")

    # Column widths
    widths = {'A': 3, 'B': 18, 'C': 14, 'D': 28, 'E': 12, 'F': 10, 'G': 15, 'H': 14, 'I': 14, 'J': 3}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header row
    ws.merge_cells('B1:I1')
    ws['B1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B1'].font = Font(name='Calibri', size=16, bold=True, color=NAVY)

    ws.merge_cells('B2:I2')
    ws['B2'] = "500HR MAGNETO INSPECTION SERVICE QUOTE"
    ws['B2'].font = header_font
    ws['B2'].fill = header_fill
    ws['B2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 30

    # Company info
    ws['B3'] = "A Hartzell Engine Technologies Company"
    ws['B3'].font = Font(size=9, italic=True, color=SKY_BLUE)

    # Quote info section
    ws['B5'] = "Quote #:"
    ws['C5'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['C5'].font = Font(bold=True, size=12, color=NAVY)

    ws['E5'] = "Date:"
    ws['F5'] = datetime.now().strftime("%m/%d/%Y")
    ws['F5'].number_format = 'MM/DD/YYYY'

    ws['B6'] = "Customer Email:"
    ws['C6'] = ""  # Input cell
    ws['C6'].fill = input_fill
    ws['C6'].border = thin_border
    ws.merge_cells('C6:E6')

    ws['G6'] = "Approval Req?"
    ws['H6'] = "No"
    ws['H6'].fill = input_fill

    # Data validation for Yes/No
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_yesno)
    dv_yesno.add(ws['H6'])

    # Magneto table header
    ws['B9'] = "Magneto P/N"
    ws['C9'] = "Serial #"
    ws['D9'] = "Description"
    ws['E9'] = "500HR Price"
    ws['F9'] = "Ovhl Req?"
    ws['G9'] = "CO Part #"
    ws['H9'] = "CO Price"
    ws['I9'] = "Line Total"

    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws[f'{col}9'].font = Font(bold=True, size=10, color=WHITE)
        ws[f'{col}9'].fill = header_fill
        ws[f'{col}9'].border = thin_border
        ws[f'{col}9'].alignment = Alignment(horizontal='center')

    # Magneto rows (6 units)
    for row in range(10, 16):
        # P/N input
        ws[f'B{row}'] = ""
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = thin_border

        # Serial input
        ws[f'C{row}'] = ""
        ws[f'C{row}'].fill = input_fill
        ws[f'C{row}'].border = thin_border

        # Description lookup - CRITICAL: Using 5H suffix (NOT -5H!)
        ws[f'D{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row}&"5H",\'Parts Master\'!$A:$H,2,FALSE),VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE)))'
        ws[f'D{row}'].border = thin_border

        # 500HR Price lookup - CRITICAL: Using 5H suffix (NOT -5H!)
        ws[f'E{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row}&"5H",\'Parts Master\'!$A:$H,4,FALSE),Config!$B$5))'
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '$#,##0.00'

        # Overhaul Required dropdown
        ws[f'F{row}'] = ""
        ws[f'F{row}'].fill = input_fill
        ws[f'F{row}'].border = thin_border
        dv_yesno.add(ws[f'F{row}'])

        # CO Part Number (auto-generated)
        ws[f'G{row}'] = f'=IF(F{row}="Yes",SUBSTITUTE(SUBSTITUTE(B{row},"5H",""),"-OH","")&"-CO","")'
        ws[f'G{row}'].border = thin_border

        # CO Price lookup
        ws[f'H{row}'] = f'=IF(G{row}="","",IFERROR(VLOOKUP(G{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        ws[f'H{row}'].border = thin_border
        ws[f'H{row}'].number_format = '$#,##0.00'

        # Line Total
        ws[f'I{row}'] = f'=IF(B{row}="","",IF(F{row}="Yes",H{row},E{row}))'
        ws[f'I{row}'].border = thin_border
        ws[f'I{row}'].number_format = '$#,##0.00'

    # Findings section header
    ws['B17'] = "INSPECTION FINDINGS"
    ws['B17'].font = Font(bold=True, size=11, color=NAVY)
    ws.merge_cells('B17:I17')

    # Findings checkboxes header
    findings = ["Bad Coil", "Bad Capacitor", "Bad Dist Block", "Bad Impulse",
                "Worn Gear", "Bad Points", "Bad Bearings", "Oil in Housing ⚠"]

    ws['B18'] = "Finding"
    for i, finding in enumerate(findings):
        col = get_column_letter(2 + i)
        ws[f'{col}18'] = finding[:12]
        ws[f'{col}18'].font = Font(size=8, bold=True)
        ws[f'{col}18'].alignment = Alignment(horizontal='center', wrap_text=True)

    # Checkbox grid (☐ unchecked, ☑ checked)
    for row in range(19, 25):  # 6 rows for 6 magnetos
        ws[f'A{row}'] = f'Unit {row-18}'
        ws[f'A{row}'].font = Font(size=8)
        for col_num in range(2, 10):  # 8 columns for 8 findings
            col = get_column_letter(col_num)
            ws[f'{col}{row}'] = "☐"
            ws[f'{col}{row}'].font = Font(size=14)
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
            ws[f'{col}{row}'].border = thin_border

    # Add-on fee calculation
    ws['B26'] = "Units with Findings:"
    ws['D26'] = '=SUMPRODUCT((COUNTIF(B19:I19,"☑")+COUNTIF(B20:I20,"☑")+COUNTIF(B21:I21,"☑")+COUNTIF(B22:I22,"☑")+COUNTIF(B23:I23,"☑")+COUNTIF(B24:I24,"☑")>0)*1)'

    ws['B27'] = "Findings Add-On ($110/unit):"
    ws['D27'] = '=D26*Config!$B$6'
    ws['D27'].number_format = '$#,##0.00'

    # Totals section
    ws['G28'] = "500HR Service Total:"
    ws['I28'] = '=SUMIF(F10:F15,"<>Yes",E10:E15)'
    ws['I28'].number_format = '$#,##0.00'
    ws['I28'].font = Font(bold=True)

    ws['G29'] = "Overhaul Total:"
    ws['I29'] = '=SUMIF(F10:F15,"Yes",H10:H15)'
    ws['I29'].number_format = '$#,##0.00'

    ws['G30'] = "GRAND TOTAL:"
    ws['H30'] = '=I28+I29+D27'
    ws['H30'].number_format = '$#,##0.00'
    ws['H30'].font = Font(bold=True, size=14, color=NAVY)
    ws['H30'].fill = total_fill
    ws['H30'].border = medium_border
    ws.merge_cells('H30:I30')

    # Notes
    ws['B32'] = "Notes:"
    ws['B32'].font = Font(bold=True)
    ws.merge_cells('B33:I35')
    ws['B33'] = ""
    ws['B33'].fill = input_fill
    ws['B33'].border = thin_border

    # Footer
    ws['B37'] = "Thank you for choosing Quality Aircraft Accessories!"
    ws['B37'].font = Font(italic=True, color='666666')

    # Print area (exclude button row)
    ws.print_area = 'A1:J37'

    return ws

def create_cust_property_sheet(wb):
    """Create Customer Property Overhaul sheet"""
    ws = wb.create_sheet("Cust Property")

    # Column widths
    widths = {'A': 8, 'B': 18, 'C': 25, 'D': 12, 'E': 12, 'F': 12, 'G': 12, 'H': 10, 'I': 12, 'J': 3}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header
    ws.merge_cells('B1:I1')
    ws['B1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B1'].font = Font(name='Calibri', size=16, bold=True, color=NAVY)

    ws.merge_cells('B2:I2')
    ws['B2'] = "CUSTOMER PROPERTY OVERHAUL QUOTE"
    ws['B2'].font = header_font
    ws['B2'].fill = header_fill
    ws['B2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 30

    ws['B3'] = "A Hartzell Engine Technologies Company"
    ws['B3'].font = Font(size=9, italic=True, color=SKY_BLUE)

    # Quote info
    ws['B5'] = "Quote #:"
    ws['C5'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['C5'].font = Font(bold=True, size=12, color=NAVY)

    ws['E5'] = "Date:"
    ws['F5'] = datetime.now().strftime("%m/%d/%Y")

    ws['H5'] = "Updated:"
    ws['I5'] = ""
    ws['I5'].fill = input_fill

    ws['B6'] = "Customer Email:"
    ws['C6'] = ""
    ws['C6'].fill = input_fill
    ws['C6'].border = thin_border
    ws.merge_cells('C6:E6')

    ws['G6'] = "Tier:"
    ws['H6'] = "Standard"
    ws['H6'].fill = input_fill

    # Tier dropdown
    dv_tier = DataValidation(type="list", formula1='"Standard,Tier 1,Tier 2,Tier 3"', allow_blank=True)
    ws.add_data_validation(dv_tier)
    dv_tier.add(ws['H6'])

    # Requirements
    ws['B7'] = "Approval?"
    ws['C7'] = "No"
    ws['D7'] = "Shipping?"
    ws['E7'] = "No"
    ws['F7'] = "Payment?"
    ws['G7'] = "No"

    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_yesno)
    for cell in ['C7', 'E7', 'G7']:
        ws[cell].fill = input_fill
        dv_yesno.add(ws[cell])

    # Units table header
    ws['B10'] = "Unit P/N (-CO)"
    ws['C10'] = "Description"
    ws['D10'] = "Type"
    ws['E10'] = "Base Labor"
    ws['F10'] = "Your Price"
    ws['G10'] = "Exchange Cap"

    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}10'].font = Font(bold=True, size=10, color=WHITE)
        ws[f'{col}10'].fill = header_fill
        ws[f'{col}10'].border = thin_border

    # Unit rows (4 units)
    for row in range(11, 15):
        ws[f'B{row}'] = ""
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = thin_border

        # Description lookup
        ws[f'C{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found"))'
        ws[f'C{row}'].border = thin_border

        # Type (In-House vs Vendor)
        ws[f'D{row}'] = f'=IF(B{row}="","",IF(COUNTIF(BOM!$A:$A,SUBSTITUTE(B{row},"-CO",""))>0,"In-House","Vendor"))'
        ws[f'D{row}'].border = thin_border

        # Base Labor (manual input)
        ws[f'E{row}'] = ""
        ws[f'E{row}'].fill = input_fill
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '$#,##0.00'

        # Your Price (with tier discount)
        ws[f'F{row}'] = f'=IF(E{row}="","",E{row}*(1-VLOOKUP($H$6,Config!$A$10:$B$13,2,FALSE)))'
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].number_format = '$#,##0.00'

        # Exchange Cap reference
        ws[f'G{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(SUBSTITUTE(B{row},"-CO","-OH"),\'Parts Master\'!$A:$H,4,FALSE),"N/A"))'
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].number_format = '$#,##0.00'

    # O&A Parts section
    ws['B16'] = "OVER & ABOVE PARTS"
    ws['B16'].font = Font(bold=True, size=11, color=NAVY)

    # O&A header
    ws['A17'] = "Unit#"
    ws['B17'] = "Part #"
    ws['C17'] = "Description"
    ws['D17'] = "Qty"
    ws['E17'] = "Unit Price"
    ws['F17'] = "Override"
    ws['G17'] = "Extended"
    ws['H17'] = "Status"

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws[f'{col}17'].font = Font(bold=True, size=9, color=WHITE)
        ws[f'{col}17'].fill = header_fill
        ws[f'{col}17'].border = thin_border

    # O&A rows (10 parts)
    for row in range(18, 28):
        ws[f'A{row}'] = ""
        ws[f'A{row}'].fill = input_fill
        ws[f'A{row}'].border = thin_border

        ws[f'B{row}'] = ""
        ws[f'B{row}'].fill = input_fill
        ws[f'B{row}'].border = thin_border

        ws[f'C{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),""))'
        ws[f'C{row}'].border = thin_border

        ws[f'D{row}'] = ""
        ws[f'D{row}'].fill = input_fill
        ws[f'D{row}'].border = thin_border

        ws[f'E{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '$#,##0.00'

        ws[f'F{row}'] = ""
        ws[f'F{row}'].fill = input_fill
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].number_format = '$#,##0.00'

        ws[f'G{row}'] = f'=IF(D{row}="","",IF(F{row}<>"",F{row}*D{row},E{row}*D{row}))'
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].number_format = '$#,##0.00'

        ws[f'H{row}'] = ""
        ws[f'H{row}'].fill = input_fill
        ws[f'H{row}'].border = thin_border

    # Totals
    ws['F29'] = "Labor Total:"
    ws['G29'] = '=SUM(F11:F14)'
    ws['G29'].number_format = '$#,##0.00'
    ws['G29'].font = Font(bold=True)

    ws['F30'] = "Parts Total:"
    ws['G30'] = '=SUM(G18:G27)'
    ws['G30'].number_format = '$#,##0.00'
    ws['G30'].font = Font(bold=True)

    ws['F31'] = "GRAND TOTAL:"
    ws['G31'] = '=G29+G30'
    ws['G31'].number_format = '$#,##0.00'
    ws['G31'].font = Font(bold=True, size=14, color=NAVY)
    ws['G31'].fill = total_fill
    ws['G31'].border = medium_border

    # Notes
    ws['B33'] = "Notes:"
    ws.merge_cells('B34:H35')
    ws['B34'] = ""
    ws['B34'].fill = input_fill
    ws['B34'].border = thin_border

    ws.print_area = 'A1:I36'

    return ws

def create_exchange_sheet(wb):
    """Create Exchange/Inventory Sale sheet"""
    ws = wb.create_sheet("Exchange")

    widths = {'A': 18, 'B': 25, 'C': 8, 'D': 6, 'E': 12, 'F': 12, 'G': 12, 'H': 12, 'I': 10, 'J': 12}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color=NAVY)

    ws.merge_cells('A2:J2')
    ws['A2'] = "EXCHANGE / INVENTORY SALE QUOTE"
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 30

    ws['A3'] = "A Hartzell Engine Technologies Company"
    ws['A3'].font = Font(size=9, italic=True, color=SKY_BLUE)

    # Quote info
    ws['A5'] = "Quote #:"
    ws['B5'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['B5'].font = Font(bold=True, size=12, color=NAVY)

    ws['D5'] = "Date:"
    ws['E5'] = datetime.now().strftime("%m/%d/%Y")

    ws['A6'] = "Customer Email:"
    ws['B6'] = ""
    ws['B6'].fill = input_fill
    ws['B6'].border = thin_border
    ws.merge_cells('B6:D6')

    ws['F6'] = "Tier:"
    ws['G6'] = "Standard"
    ws['G6'].fill = input_fill

    dv_tier = DataValidation(type="list", formula1='"Standard,Tier 1,Tier 2,Tier 3"', allow_blank=True)
    ws.add_data_validation(dv_tier)
    dv_tier.add(ws['G6'])

    # Requirements row
    ws['A7'] = "Approval?"
    ws['B7'] = "No"
    ws['C7'] = "Shipping?"
    ws['D7'] = "No"
    ws['E7'] = "Payment?"
    ws['F7'] = "No"
    ws['G7'] = "Core In Hand?"
    ws['H7'] = "No"

    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_yesno)
    for cell in ['B7', 'D7', 'F7', 'H7']:
        ws[cell].fill = input_fill
        dv_yesno.add(ws[cell])

    # Line items header
    headers = ['Part #', 'Description', 'Cond', 'Qty', 'List Price', 'Your Price', 'Extended', 'Core Charge', 'Stock', 'Lead Time']
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}10'] = header
        ws[f'{col}10'].font = Font(bold=True, size=10, color=WHITE)
        ws[f'{col}10'].fill = header_fill
        ws[f'{col}10'].border = thin_border

    # Line item rows (12 items)
    for row in range(11, 23):
        # Part #
        ws[f'A{row}'] = ""
        ws[f'A{row}'].fill = input_fill
        ws[f'A{row}'].border = thin_border

        # Description lookup
        ws[f'B{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found"))'
        ws[f'B{row}'].border = thin_border

        # Condition lookup
        ws[f'C{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,3,FALSE),""))'
        ws[f'C{row}'].border = thin_border

        # Qty
        ws[f'D{row}'] = ""
        ws[f'D{row}'].fill = input_fill
        ws[f'D{row}'].border = thin_border

        # List Price
        ws[f'E{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        ws[f'E{row}'].border = thin_border
        ws[f'E{row}'].number_format = '$#,##0.00'

        # Your Price (with tier discount)
        ws[f'F{row}'] = f'=IF(E{row}="","",E{row}*(1-VLOOKUP($G$6,Config!$A$10:$B$13,2,FALSE)))'
        ws[f'F{row}'].border = thin_border
        ws[f'F{row}'].number_format = '$#,##0.00'

        # Extended
        ws[f'G{row}'] = f'=IF(D{row}="","",F{row}*D{row})'
        ws[f'G{row}'].border = thin_border
        ws[f'G{row}'].number_format = '$#,##0.00'

        # Core Charge (waived if Core In Hand = Yes)
        ws[f'H{row}'] = f'=IF(A{row}="","",IF($H$7="Yes",0,IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,5,FALSE),0)*D{row}))'
        ws[f'H{row}'].border = thin_border
        ws[f'H{row}'].number_format = '$#,##0.00'

        # Stock
        ws[f'I{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,6,FALSE),""))'
        ws[f'I{row}'].border = thin_border

        # Lead Time
        ws[f'J{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,7,FALSE),""))'
        ws[f'J{row}'].border = thin_border

    # Totals
    ws['A24'] = "Subtotal:"
    ws['B24'] = '=SUM(G11:G22)'
    ws['B24'].number_format = '$#,##0.00'
    ws['B24'].font = Font(bold=True)

    ws['C24'] = "Core Total:"
    ws['D24'] = '=SUM(H11:H22)'
    ws['D24'].number_format = '$#,##0.00'
    if ws['H7'].value == "Yes":
        ws['E24'] = "(WAIVED)"
        ws['E24'].font = Font(color='008000')

    ws['F24'] = "Your Savings:"
    ws['G24'] = '=SUM(E11:E22)-SUM(F11:F22)'
    ws['G24'].number_format = '$#,##0.00'
    ws['G24'].font = Font(color='008000', bold=True)

    ws['D25'] = "GRAND TOTAL:"
    ws['E25'] = '=B24+IF(H7="Yes",0,D24)'
    ws['E25'].number_format = '$#,##0.00'
    ws['E25'].font = Font(bold=True, size=14, color=NAVY)
    ws['E25'].fill = total_fill
    ws['E25'].border = medium_border
    ws.merge_cells('E25:F25')

    # Notes
    ws['A27'] = "Notes:"
    ws.merge_cells('A28:J29')
    ws['A28'] = ""
    ws['A28'].fill = input_fill
    ws['A28'].border = thin_border

    ws.print_area = 'A1:J30'

    return ws

def create_parts_master_sheet(wb):
    """Create Parts Master database sheet"""
    ws = wb.create_sheet("Parts Master")

    widths = {'A': 18, 'B': 35, 'C': 8, 'D': 12, 'E': 12, 'F': 8, 'G': 12, 'H': 30}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header
    headers = ['Part Number', 'Description', 'Cond', 'List Price', 'Core Charge', 'Stock', 'Lead Time', 'Notes']
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}1'] = header
        ws[f'{col}1'].font = Font(bold=True, color=WHITE)
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].border = thin_border

    # Sample data with CORRECT 5H suffix (NOT -5H!)
    parts_data = [
        # S-20/200 Magneto
        ['10-382320', 'S-20/200 Magneto', 'Base', 0, 0, 0, '', 'Base reference'],
        ['10-3823205H', 'S-20/200 Magneto - 500HR Inspection', 'Svc', 575, 0, 0, '3-5 Days', '500HR service price'],
        ['10-382320-CO', 'S-20/200 Magneto - Customer Overhaul', 'CO', 895, 0, 0, '7-10 Days', 'Cust prop overhaul'],
        ['10-382320-OH', 'S-20/200 Magneto - OH Exchange', 'OH', 2495, 1000, 5, '5-7 Days', 'Exchange unit'],

        # S-1200 Magneto
        ['10-320680', 'S-1200 Magneto', 'Base', 0, 0, 0, '', 'Base reference'],
        ['10-3206805H', 'S-1200 Magneto - 500HR Inspection', 'Svc', 495, 0, 0, '3-5 Days', '500HR service price'],
        ['10-320680-CO', 'S-1200 Magneto - Customer Overhaul', 'CO', 750, 0, 0, '7-10 Days', 'Cust prop overhaul'],
        ['10-320680-OH', 'S-1200 Magneto - OH Exchange', 'OH', 2095, 800, 3, '5-7 Days', 'Exchange unit'],

        # D-2000/D-3000 Magneto
        ['10-391088', 'D-2000/D-3000 Magneto', 'Base', 0, 0, 0, '', 'Base reference'],
        ['10-3910885H', 'D-2000/D-3000 - 500HR Inspection', 'Svc', 695, 0, 0, '3-5 Days', '500HR service price'],
        ['10-391088-CO', 'D-2000/D-3000 - Customer Overhaul', 'CO', 1150, 0, 0, '7-10 Days', 'Cust prop overhaul'],
        ['10-391088-OH', 'D-2000/D-3000 - OH Exchange', 'OH', 2850, 1100, 2, '5-7 Days', 'Exchange unit'],

        # Bendix Magneto
        ['10-163020', 'Bendix S-20 Magneto', 'Base', 0, 0, 0, '', 'Base reference'],
        ['10-1630205H', 'Bendix S-20 - 500HR Inspection', 'Svc', 625, 0, 0, '3-5 Days', '500HR service price'],
        ['10-163020-CO', 'Bendix S-20 - Customer Overhaul', 'CO', 950, 0, 0, '7-10 Days', 'Cust prop overhaul'],
        ['10-163020-OH', 'Bendix S-20 - OH Exchange', 'OH', 2650, 950, 4, '5-7 Days', 'Exchange unit'],

        # Component parts
        ['LW-15473', 'Impulse Coupling Assembly', 'New', 425, 0, 12, 'In Stock', 'Common replacement'],
        ['10-357170', 'Coil Assembly', 'New', 185, 0, 20, 'In Stock', 'Common replacement'],
        ['10-357210', 'Capacitor', 'New', 45, 0, 50, 'In Stock', 'Common replacement'],
        ['10-51360', 'Distributor Block', 'New', 125, 0, 15, 'In Stock', 'Common replacement'],
        ['10-51350', 'Points Assembly', 'New', 85, 0, 30, 'In Stock', 'Common replacement'],
        ['10-51340', 'Bearing Set', 'New', 95, 0, 25, 'In Stock', 'Common replacement'],
        ['10-51330', 'Gear Assembly', 'New', 165, 0, 8, 'In Stock', 'Common replacement'],

        # Governor
        ['646942-CO', 'Governor - Customer Overhaul', 'CO', 1500, 0, 0, '14-21 Days', 'Vendor overhaul'],
        ['646942-OH', 'Governor - OH Exchange', 'OH', 3450, 1500, 2, '5-7 Days', 'Exchange unit'],

        # Turbocharger
        ['478615-CO', 'Turbocharger - Customer Overhaul', 'CO', 2800, 0, 0, '21-30 Days', 'Vendor overhaul'],
        ['478615-OH', 'Turbocharger - OH Exchange', 'OH', 5800, 2500, 1, '5-7 Days', 'Exchange unit'],
    ]

    for row_num, row_data in enumerate(parts_data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            if col_num in [4, 5]:  # Price columns
                cell.number_format = '$#,##0.00'

    return ws

def create_bom_sheet(wb):
    """Create Bill of Materials sheet"""
    ws = wb.create_sheet("BOM")

    headers = ['Parent P/N', 'Parent Description', 'Component P/N', 'Component Description', 'Qty Per']
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}1'] = header
        ws[f'{col}1'].font = Font(bold=True, color=WHITE)
        ws[f'{col}1'].fill = header_fill
        ws.column_dimensions[col].width = 25

    # Sample BOM data (determines In-House vs Vendor)
    bom_data = [
        ['10-382320', 'S-20/200 Magneto', '10-357170', 'Coil Assembly', 1],
        ['10-382320', 'S-20/200 Magneto', '10-357210', 'Capacitor', 1],
        ['10-382320', 'S-20/200 Magneto', '10-51360', 'Distributor Block', 1],
        ['10-382320', 'S-20/200 Magneto', 'LW-15473', 'Impulse Coupling', 1],
        ['10-320680', 'S-1200 Magneto', '10-357170', 'Coil Assembly', 1],
        ['10-320680', 'S-1200 Magneto', '10-357210', 'Capacitor', 1],
        ['10-391088', 'D-2000/D-3000 Magneto', '10-357170', 'Coil Assembly', 1],
        ['10-391088', 'D-2000/D-3000 Magneto', '10-51360', 'Distributor Block', 1],
    ]

    for row_num, row_data in enumerate(bom_data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value).border = thin_border

    return ws

def create_config_sheet(wb):
    """Create Config settings sheet"""
    ws = wb.create_sheet("Config")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 30

    # Header
    ws['A1'] = "SYSTEM CONFIGURATION"
    ws['A1'].font = Font(bold=True, size=14, color=NAVY)

    # Settings
    settings = [
        ('Quote Prefix', 'QAA'),
        ('Next Quote #', 1),
        ('Quote Valid Days', 30),
        ('500HR Default Fee', 550),
        ('500HR Findings Add-On', 110),
    ]

    for i, (name, value) in enumerate(settings, start=2):
        ws[f'A{i}'] = name
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
        ws[f'B{i}'].fill = input_fill
        ws[f'B{i}'].border = thin_border

    # Notes
    ws['C3'] = "Auto-increments after each email sent"
    ws['C5'] = "Fallback if 5H variant not found"
    ws['C6'] = "Flat fee per unit with ANY findings"

    # Tier discounts
    ws['A9'] = "TIER DISCOUNTS"
    ws['A9'].font = Font(bold=True, size=12, color=NAVY)

    tiers = [
        ('Standard', 0),
        ('Tier 1', 0.07),
        ('Tier 2', 0.10),
        ('Tier 3', 0.13),
    ]

    for i, (tier, discount) in enumerate(tiers, start=10):
        ws[f'A{i}'] = tier
        ws[f'B{i}'] = discount
        ws[f'B{i}'].number_format = '0%'
        ws[f'B{i}'].border = thin_border

    return ws

def create_quote_log_sheet(wb):
    """Create Quote Log sheet"""
    ws = wb.create_sheet("Quote Log")

    headers = ['Quote #', 'Date Sent', 'Time', 'Customer Email', 'Type', 'Units',
               'Total', 'Core', 'Status', 'Subject', 'Body Preview']

    widths = [12, 12, 10, 30, 18, 8, 12, 12, 10, 40, 50]

    for i, (header, width) in enumerate(zip(headers, widths)):
        col = get_column_letter(i + 1)
        ws[f'{col}1'] = header
        ws[f'{col}1'].font = Font(bold=True, color=WHITE)
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].border = thin_border
        ws.column_dimensions[col].width = width

    return ws

def create_email_log_sheet(wb):
    """Create Email Log sheet with full email body storage"""
    ws = wb.create_sheet("Email Log")

    headers = ['Quote #', 'Date Sent', 'Time', 'Customer Email', 'Subject', 'Full Email Body']
    widths = [12, 12, 10, 30, 50, 100]

    for i, (header, width) in enumerate(zip(headers, widths)):
        col = get_column_letter(i + 1)
        ws[f'{col}1'] = header
        ws[f'{col}1'].font = Font(bold=True, color=WHITE)
        ws[f'{col}1'].fill = header_fill
        ws[f'{col}1'].border = thin_border
        ws.column_dimensions[col].width = width

    return ws

def create_inventory_sheet(wb):
    """Create Inventory sheet for MAPICS data"""
    ws = wb.create_sheet("Inventory")

    headers = ['Part Number', 'Description', 'On Hand', 'Allocated', 'Available', 'Location', 'Last Updated']

    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}1'] = header
        ws[f'{col}1'].font = Font(bold=True, color=WHITE)
        ws[f'{col}1'].fill = header_fill
        ws.column_dimensions[col].width = 18

    return ws

def create_vba_module():
    """Generate the corrected VBA module code with 5H suffix (NOT -5H)"""

    vba_code = '''Attribute VB_Name = "QAAQuoteSystem"
' ╔══════════════════════════════════════════════════════════════════════════════╗
' ║     QAA QUOTE SYSTEM v12.0 - VBA MODULE                                      ║
' ║     Quality Aircraft Accessories - A Hartzell Engine Technologies Company    ║
' ║     CORRECTED: Uses 5H suffix (NOT -5H) for 500HR pricing                    ║
' ║     Full Email Logging: Customer email, subject, FULL body, date/time        ║
' ╚══════════════════════════════════════════════════════════════════════════════╝
'
' SETUP INSTRUCTIONS:
' 1. Save workbook as .xlsm (macro-enabled)
' 2. Press Alt+F11 to open VBA Editor
' 3. Insert > Module
' 4. Paste this entire code
' 5. Close VBA Editor
' 6. Right-click each button image > Assign Macro > Select the appropriate macro

Option Explicit

' ═══════════════════════════════════════════════════════════════════════════════
' NAVIGATION MACROS
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub GoMenu()
    Application.ScreenUpdating = False
    ThisWorkbook.Sheets("Menu").Activate
    Application.ScreenUpdating = True
End Sub

Public Sub Go500HR()
    Application.ScreenUpdating = False
    ThisWorkbook.Sheets("500HR").Activate
    ThisWorkbook.Sheets("500HR").Range("B6").Select
    Application.ScreenUpdating = True
End Sub

Public Sub GoCustProp()
    Application.ScreenUpdating = False
    ThisWorkbook.Sheets("Cust Property").Activate
    ThisWorkbook.Sheets("Cust Property").Range("C6").Select
    Application.ScreenUpdating = True
End Sub

Public Sub GoExchange()
    Application.ScreenUpdating = False
    ThisWorkbook.Sheets("Exchange").Activate
    ThisWorkbook.Sheets("Exchange").Range("B6").Select
    Application.ScreenUpdating = True
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' PDF GENERATION
' ═══════════════════════════════════════════════════════════════════════════════

Private Function GetPDFPath(quoteNum As String) As String
    Dim folder As String
    folder = ThisWorkbook.Path
    If folder = "" Then folder = Environ("USERPROFILE") & "\\Documents"
    GetPDFPath = folder & "\\Quote_" & Replace(Replace(quoteNum, "-", "_"), "/", "_") & "_" & Format(Date, "YYYYMMDD") & ".pdf"
End Function

Public Sub GenPDF500HR()
    Dim ws As Worksheet
    Dim pdfPath As String

    Set ws = ThisWorkbook.Sheets("500HR")

    If Trim(ws.Range("C6").Value) = "" Then
        MsgBox "Please enter customer email first.", vbExclamation, "QAA Quote System"
        ws.Range("C6").Select
        Exit Sub
    End If

    pdfPath = GetPDFPath(ws.Range("C5").Value)

    Application.ScreenUpdating = False

    With ws.PageSetup
        .PrintArea = "$A$1:$J$37"
        .Orientation = xlPortrait
        .FitToPagesWide = 1
        .FitToPagesTall = 1
        .LeftMargin = Application.InchesToPoints(0.4)
        .RightMargin = Application.InchesToPoints(0.4)
        .TopMargin = Application.InchesToPoints(0.4)
        .BottomMargin = Application.InchesToPoints(0.4)
        .CenterHorizontally = True
    End With

    On Error GoTo PDFError
    ws.ExportAsFixedFormat xlTypePDF, pdfPath, xlQualityStandard, False, False, , , False

    Application.ScreenUpdating = True
    MsgBox "PDF Created Successfully!" & vbCrLf & vbCrLf & pdfPath, vbInformation, "QAA Quote System"

    If MsgBox("Create email with PDF attached?", vbYesNo + vbQuestion, "QAA Quote System") = vbYes Then
        Email500HR pdfPath
    End If
    Exit Sub

PDFError:
    Application.ScreenUpdating = True
    MsgBox "PDF Error: " & Err.Description, vbCritical, "QAA Quote System"
End Sub

Public Sub GenPDFCustProp()
    Dim ws As Worksheet
    Dim pdfPath As String

    Set ws = ThisWorkbook.Sheets("Cust Property")

    If Trim(ws.Range("C6").Value) = "" Then
        MsgBox "Please enter customer email first.", vbExclamation, "QAA Quote System"
        ws.Range("C6").Select
        Exit Sub
    End If

    pdfPath = GetPDFPath(ws.Range("C5").Value)

    Application.ScreenUpdating = False

    With ws.PageSetup
        .PrintArea = "$A$1:$I$36"
        .Orientation = xlPortrait
        .FitToPagesWide = 1
        .FitToPagesTall = 1
        .LeftMargin = Application.InchesToPoints(0.4)
        .RightMargin = Application.InchesToPoints(0.4)
        .TopMargin = Application.InchesToPoints(0.4)
        .BottomMargin = Application.InchesToPoints(0.4)
        .CenterHorizontally = True
    End With

    On Error GoTo PDFError
    ws.ExportAsFixedFormat xlTypePDF, pdfPath, xlQualityStandard, False, False, , , False

    Application.ScreenUpdating = True
    MsgBox "PDF Created Successfully!" & vbCrLf & vbCrLf & pdfPath, vbInformation, "QAA Quote System"

    If MsgBox("Create email with PDF attached?", vbYesNo + vbQuestion, "QAA Quote System") = vbYes Then
        EmailCustProp pdfPath
    End If
    Exit Sub

PDFError:
    Application.ScreenUpdating = True
    MsgBox "PDF Error: " & Err.Description, vbCritical, "QAA Quote System"
End Sub

Public Sub GenPDFExchange()
    Dim ws As Worksheet
    Dim pdfPath As String

    Set ws = ThisWorkbook.Sheets("Exchange")

    If Trim(ws.Range("B6").Value) = "" Then
        MsgBox "Please enter customer email first.", vbExclamation, "QAA Quote System"
        ws.Range("B6").Select
        Exit Sub
    End If

    pdfPath = GetPDFPath(ws.Range("B5").Value)

    Application.ScreenUpdating = False

    With ws.PageSetup
        .PrintArea = "$A$1:$J$30"
        .Orientation = xlLandscape
        .FitToPagesWide = 1
        .FitToPagesTall = 1
        .LeftMargin = Application.InchesToPoints(0.4)
        .RightMargin = Application.InchesToPoints(0.4)
        .TopMargin = Application.InchesToPoints(0.4)
        .BottomMargin = Application.InchesToPoints(0.4)
        .CenterHorizontally = True
    End With

    On Error GoTo PDFError
    ws.ExportAsFixedFormat xlTypePDF, pdfPath, xlQualityStandard, False, False, , , False

    Application.ScreenUpdating = True
    MsgBox "PDF Created Successfully!" & vbCrLf & vbCrLf & pdfPath, vbInformation, "QAA Quote System"

    If MsgBox("Create email with PDF attached?", vbYesNo + vbQuestion, "QAA Quote System") = vbYes Then
        EmailExchange pdfPath
    End If
    Exit Sub

PDFError:
    Application.ScreenUpdating = True
    MsgBox "PDF Error: " & Err.Description, vbCritical, "QAA Quote System"
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' EMAIL GENERATION WITH COMPLETE LOGGING
' Logs: Customer email, Subject, FULL body text, Date, Time
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub Email500HR(Optional pdfPath As String = "")
    Dim OutApp As Object, OutMail As Object
    Dim ws As Worksheet
    Dim emailTo As String, subj As String, body As String
    Dim quoteNum As String, totalAmt As Double

    Set ws = ThisWorkbook.Sheets("500HR")
    emailTo = Trim(ws.Range("C6").Value)
    quoteNum = ws.Range("C5").Value
    totalAmt = ws.Range("H30").Value

    If emailTo = "" Then
        MsgBox "Please enter customer email address.", vbExclamation, "QAA Quote System"
        ws.Range("C6").Select
        Exit Sub
    End If

    ' Build professional subject line
    subj = "Quote " & quoteNum & " - 500HR Magneto Inspection Service - Quality Aircraft Accessories"

    ' Build professional email body
    body = "Dear Valued Customer," & vbCrLf & vbCrLf
    body = body & "Thank you for choosing Quality Aircraft Accessories for your magneto inspection needs." & vbCrLf & vbCrLf
    body = body & "Please find attached your quote for 500HR Magneto Inspection Service." & vbCrLf & vbCrLf
    body = body & String(54, "=") & vbCrLf
    body = body & "                    QUOTE SUMMARY" & vbCrLf
    body = body & String(54, "=") & vbCrLf
    body = body & "Quote #:        " & quoteNum & vbCrLf
    body = body & "Date:           " & Format(Date, "MMMM d, yyyy") & vbCrLf
    body = body & "Valid:          30 Days" & vbCrLf & vbCrLf
    body = body & "TOTAL:          " & Format(totalAmt, "$#,##0.00") & vbCrLf
    body = body & String(54, "=") & vbCrLf & vbCrLf

    ' Check for units requiring overhaul
    If Application.WorksheetFunction.CountIf(ws.Range("F10:F15"), "Yes") > 0 Then
        body = body & "NOTE: Some units require overhaul based on inspection findings." & vbCrLf
        body = body & "Please see attached quote for complete details." & vbCrLf & vbCrLf
    End If

    ' Check approval requirement
    If ws.Range("H6").Value = "Yes" Then
        body = body & "** APPROVAL REQUIRED **" & vbCrLf
        body = body & "Please reply with your approval to proceed with service." & vbCrLf & vbCrLf
    End If

    body = body & "If you have any questions, please don't hesitate to contact us." & vbCrLf & vbCrLf
    body = body & "Best regards," & vbCrLf & vbCrLf
    body = body & "Quality Aircraft Accessories" & vbCrLf
    body = body & "A Hartzell Engine Technologies Company" & vbCrLf
    body = body & "5746 E Apache St, Tulsa, OK 74115" & vbCrLf
    body = body & "918-835-6948 | sales@qaainc.com | qaa.com"

    On Error GoTo EmailError
    Set OutApp = CreateObject("Outlook.Application")
    Set OutMail = OutApp.CreateItem(0)

    With OutMail
        .To = emailTo
        .Subject = subj
        .body = body
        If pdfPath <> "" Then
            If Dir(pdfPath) <> "" Then .Attachments.Add pdfPath
        End If
        .Display
    End With

    ' LOG THE EMAIL WITH FULL DETAILS
    LogEmail quoteNum, emailTo, subj, body, "500HR Service", Count500HRUnits(), totalAmt, 0

    MsgBox "Email created and logged successfully!", vbInformation, "QAA Quote System"
    Exit Sub

EmailError:
    MsgBox "Email Error: " & Err.Description & vbCrLf & vbCrLf & _
           "Make sure Microsoft Outlook is installed.", vbCritical, "QAA Quote System"
End Sub

Public Sub EmailCustProp(Optional pdfPath As String = "")
    Dim OutApp As Object, OutMail As Object
    Dim ws As Worksheet
    Dim emailTo As String, subj As String, body As String
    Dim quoteNum As String, totalAmt As Double

    Set ws = ThisWorkbook.Sheets("Cust Property")
    emailTo = Trim(ws.Range("C6").Value)
    quoteNum = ws.Range("C5").Value
    totalAmt = ws.Range("G31").Value

    If emailTo = "" Then
        MsgBox "Please enter customer email address.", vbExclamation, "QAA Quote System"
        ws.Range("C6").Select
        Exit Sub
    End If

    subj = "Quote " & quoteNum & " - Customer Property Overhaul - Quality Aircraft Accessories"

    body = "Dear Valued Customer," & vbCrLf & vbCrLf
    body = body & "Thank you for choosing Quality Aircraft Accessories for your overhaul needs." & vbCrLf & vbCrLf
    body = body & "Please find attached your quote for Customer Property Overhaul services." & vbCrLf & vbCrLf
    body = body & String(54, "=") & vbCrLf
    body = body & "                    QUOTE SUMMARY" & vbCrLf
    body = body & String(54, "=") & vbCrLf
    body = body & "Quote #:        " & quoteNum & vbCrLf
    body = body & "Date:           " & Format(Date, "MMMM d, yyyy") & vbCrLf
    body = body & "Valid:          30 Days" & vbCrLf & vbCrLf
    body = body & "TOTAL:          " & Format(totalAmt, "$#,##0.00") & vbCrLf
    body = body & String(54, "=") & vbCrLf & vbCrLf

    ' Requirements section
    Dim req As String: req = ""
    If ws.Range("C7").Value = "Yes" Then req = req & "  - Approval to proceed" & vbCrLf
    If ws.Range("E7").Value = "Yes" Then req = req & "  - Shipping information" & vbCrLf
    If ws.Range("G7").Value = "Yes" Then req = req & "  - Payment method/terms" & vbCrLf

    If req <> "" Then
        body = body & "TO PROCEED, WE NEED:" & vbCrLf & req & vbCrLf
    End If

    body = body & "If you have any questions, please don't hesitate to contact us." & vbCrLf & vbCrLf
    body = body & "Best regards," & vbCrLf & vbCrLf
    body = body & "Quality Aircraft Accessories" & vbCrLf
    body = body & "A Hartzell Engine Technologies Company" & vbCrLf
    body = body & "5746 E Apache St, Tulsa, OK 74115" & vbCrLf
    body = body & "918-835-6948 | sales@qaainc.com | qaa.com"

    On Error GoTo EmailError
    Set OutApp = CreateObject("Outlook.Application")
    Set OutMail = OutApp.CreateItem(0)

    With OutMail
        .To = emailTo
        .Subject = subj
        .body = body
        If pdfPath <> "" Then
            If Dir(pdfPath) <> "" Then .Attachments.Add pdfPath
        End If
        .Display
    End With

    LogEmail quoteNum, emailTo, subj, body, "Customer Property", CountCPUnits(), totalAmt, 0

    MsgBox "Email created and logged successfully!", vbInformation, "QAA Quote System"
    Exit Sub

EmailError:
    MsgBox "Email Error: " & Err.Description, vbCritical, "QAA Quote System"
End Sub

Public Sub EmailExchange(Optional pdfPath As String = "")
    Dim OutApp As Object, OutMail As Object
    Dim ws As Worksheet
    Dim emailTo As String, subj As String, body As String
    Dim quoteNum As String, totalAmt As Double, coreAmt As Double, savings As Double

    Set ws = ThisWorkbook.Sheets("Exchange")
    emailTo = Trim(ws.Range("B6").Value)
    quoteNum = ws.Range("B5").Value
    totalAmt = ws.Range("E25").Value
    coreAmt = ws.Range("D24").Value
    savings = ws.Range("G24").Value

    If emailTo = "" Then
        MsgBox "Please enter customer email address.", vbExclamation, "QAA Quote System"
        ws.Range("B6").Select
        Exit Sub
    End If

    subj = "Quote " & quoteNum & " - Exchange/Sale Quote - Quality Aircraft Accessories"

    body = "Dear Valued Customer," & vbCrLf & vbCrLf
    body = body & "Thank you for your interest in Quality Aircraft Accessories." & vbCrLf & vbCrLf
    body = body & "Please find attached your quote." & vbCrLf & vbCrLf
    body = body & String(54, "=") & vbCrLf
    body = body & "                    QUOTE SUMMARY" & vbCrLf
    body = body & String(54, "=") & vbCrLf
    body = body & "Quote #:        " & quoteNum & vbCrLf
    body = body & "Date:           " & Format(Date, "MMMM d, yyyy") & vbCrLf
    body = body & "Valid:          30 Days" & vbCrLf & vbCrLf
    body = body & "Subtotal:       " & Format(ws.Range("B24").Value, "$#,##0.00") & vbCrLf
    body = body & "Core Charges:   " & Format(coreAmt, "$#,##0.00")
    If ws.Range("H7").Value = "Yes" Then body = body & " (WAIVED - Core In Hand)"
    body = body & vbCrLf
    body = body & String(54, "-") & vbCrLf
    body = body & "TOTAL:          " & Format(totalAmt, "$#,##0.00") & vbCrLf
    body = body & String(54, "=") & vbCrLf & vbCrLf

    If savings > 0 Then
        body = body & "YOUR SAVINGS:   " & Format(savings, "$#,##0.00") & vbCrLf & vbCrLf
    End If

    ' Requirements
    Dim req As String: req = ""
    If ws.Range("B7").Value = "Yes" Then req = req & "  - Approval to proceed" & vbCrLf
    If ws.Range("D7").Value = "Yes" Then req = req & "  - Shipping information" & vbCrLf
    If ws.Range("F7").Value = "Yes" Then req = req & "  - Payment method/terms" & vbCrLf

    If req <> "" Then
        body = body & "TO PROCEED, WE NEED:" & vbCrLf & req & vbCrLf
    End If

    If ws.Range("H7").Value <> "Yes" And coreAmt > 0 Then
        body = body & "CORE RETURN POLICY:" & vbCrLf
        body = body & "Core charges are refunded upon receipt of serviceable core" & vbCrLf
        body = body & "within 30 days of shipment." & vbCrLf & vbCrLf
    End If

    body = body & "If you have any questions, please don't hesitate to contact us." & vbCrLf & vbCrLf
    body = body & "Best regards," & vbCrLf & vbCrLf
    body = body & "Quality Aircraft Accessories" & vbCrLf
    body = body & "A Hartzell Engine Technologies Company" & vbCrLf
    body = body & "5746 E Apache St, Tulsa, OK 74115" & vbCrLf
    body = body & "918-835-6948 | sales@qaainc.com | qaa.com"

    On Error GoTo EmailError
    Set OutApp = CreateObject("Outlook.Application")
    Set OutMail = OutApp.CreateItem(0)

    With OutMail
        .To = emailTo
        .Subject = subj
        .body = body
        If pdfPath <> "" Then
            If Dir(pdfPath) <> "" Then .Attachments.Add pdfPath
        End If
        .Display
    End With

    LogEmail quoteNum, emailTo, subj, body, "Exchange/Sale", CountExItems(), totalAmt, coreAmt

    MsgBox "Email created and logged successfully!", vbInformation, "QAA Quote System"
    Exit Sub

EmailError:
    MsgBox "Email Error: " & Err.Description, vbCritical, "QAA Quote System"
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' COMPLETE EMAIL LOGGING
' Logs to BOTH Quote Log AND Email Log with FULL email body
' ═══════════════════════════════════════════════════════════════════════════════

Private Sub LogEmail(quoteNum As String, emailTo As String, subj As String, body As String, _
                     qType As String, units As Integer, total As Double, core As Double)

    Dim wsQuoteLog As Worksheet, wsEmailLog As Worksheet, wsConfig As Worksheet
    Dim nr As Long
    Dim dateSent As Date, timeSent As String

    On Error GoTo LogError

    Set wsQuoteLog = ThisWorkbook.Sheets("Quote Log")
    Set wsEmailLog = ThisWorkbook.Sheets("Email Log")
    Set wsConfig = ThisWorkbook.Sheets("Config")

    dateSent = Date
    timeSent = Format(Now, "HH:MM:SS AM/PM")

    ' ═══════════════════════════════════════════════════════════════════════════
    ' LOG TO QUOTE LOG (summary with truncated body)
    ' ═══════════════════════════════════════════════════════════════════════════
    nr = wsQuoteLog.Cells(wsQuoteLog.Rows.Count, 1).End(xlUp).Row + 1

    wsQuoteLog.Cells(nr, 1).Value = quoteNum           ' A: Quote #
    wsQuoteLog.Cells(nr, 2).Value = dateSent           ' B: Date Sent
    wsQuoteLog.Cells(nr, 3).Value = timeSent           ' C: Time Sent
    wsQuoteLog.Cells(nr, 4).Value = emailTo            ' D: Customer Email
    wsQuoteLog.Cells(nr, 5).Value = qType              ' E: Quote Type
    wsQuoteLog.Cells(nr, 6).Value = units              ' F: Units/Items
    wsQuoteLog.Cells(nr, 7).Value = total              ' G: Total Amount
    wsQuoteLog.Cells(nr, 7).NumberFormat = "$#,##0.00"
    wsQuoteLog.Cells(nr, 8).Value = core               ' H: Core Charges
    wsQuoteLog.Cells(nr, 8).NumberFormat = "$#,##0.00"
    wsQuoteLog.Cells(nr, 9).Value = "Sent"             ' I: Status
    wsQuoteLog.Cells(nr, 10).Value = subj              ' J: Email Subject
    wsQuoteLog.Cells(nr, 11).Value = Left(body, 500)   ' K: Body Preview (truncated)

    ' ═══════════════════════════════════════════════════════════════════════════
    ' LOG TO EMAIL LOG (FULL email body - NO truncation)
    ' ═══════════════════════════════════════════════════════════════════════════
    nr = wsEmailLog.Cells(wsEmailLog.Rows.Count, 1).End(xlUp).Row + 1

    wsEmailLog.Cells(nr, 1).Value = quoteNum           ' A: Quote #
    wsEmailLog.Cells(nr, 2).Value = dateSent           ' B: Date Sent
    wsEmailLog.Cells(nr, 3).Value = timeSent           ' C: Time Sent
    wsEmailLog.Cells(nr, 4).Value = emailTo            ' D: Customer Email
    wsEmailLog.Cells(nr, 5).Value = subj               ' E: Subject
    wsEmailLog.Cells(nr, 6).Value = body               ' F: FULL Email Body (complete text)

    ' Auto-wrap the body cell for readability
    wsEmailLog.Cells(nr, 6).WrapText = True
    wsEmailLog.Rows(nr).RowHeight = 100

    ' ═══════════════════════════════════════════════════════════════════════════
    ' INCREMENT QUOTE NUMBER
    ' ═══════════════════════════════════════════════════════════════════════════
    wsConfig.Range("B3").Value = wsConfig.Range("B3").Value + 1

    Exit Sub

LogError:
    MsgBox "Logging Error: " & Err.Description, vbExclamation, "QAA Quote System"
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' CLEAR FORM MACROS
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub Clear500HR()
    Dim ws As Worksheet
    Dim r As Integer, c As Integer

    Set ws = ThisWorkbook.Sheets("500HR")

    If MsgBox("Clear the 500HR quote form?" & vbCrLf & vbCrLf & _
              "This will reset all entries.", vbYesNo + vbQuestion, "QAA Quote System") = vbNo Then
        Exit Sub
    End If

    Application.ScreenUpdating = False

    ' Clear customer email
    ws.Range("C6").ClearContents

    ' Reset approval
    ws.Range("H6").Value = "No"

    ' Clear magneto rows
    For r = 10 To 15
        ws.Cells(r, 2).ClearContents  ' P/N
        ws.Cells(r, 3).ClearContents  ' S/N
        ws.Cells(r, 6).ClearContents  ' Overhaul Required
    Next r

    ' Reset findings checkboxes to unchecked
    For r = 19 To 24
        For c = 2 To 9
            ws.Cells(r, c).Value = ChrW(9744)  ' ☐
        Next c
    Next r

    ' Clear notes
    ws.Range("B33").ClearContents

    Application.ScreenUpdating = True
    MsgBox "Form cleared!", vbInformation, "QAA Quote System"
End Sub

Public Sub ClearCustProp()
    Dim ws As Worksheet
    Dim r As Integer

    Set ws = ThisWorkbook.Sheets("Cust Property")

    If MsgBox("Clear the Customer Property quote form?" & vbCrLf & vbCrLf & _
              "This will reset all entries.", vbYesNo + vbQuestion, "QAA Quote System") = vbNo Then
        Exit Sub
    End If

    Application.ScreenUpdating = False

    ws.Range("C6").ClearContents      ' Customer email
    ws.Range("I5").ClearContents      ' Updated date
    ws.Range("H6").Value = "Standard" ' Tier
    ws.Range("C7").Value = "No"       ' Approval
    ws.Range("E7").Value = "No"       ' Shipping
    ws.Range("G7").Value = "No"       ' Payment

    ' Clear unit rows
    For r = 11 To 14
        ws.Cells(r, 2).ClearContents  ' P/N
        ws.Cells(r, 5).ClearContents  ' Base Labor
    Next r

    ' Clear O&A rows
    For r = 18 To 27
        ws.Cells(r, 1).ClearContents  ' Unit#
        ws.Cells(r, 2).ClearContents  ' Part#
        ws.Cells(r, 4).ClearContents  ' Qty
        ws.Cells(r, 6).ClearContents  ' Override
        ws.Cells(r, 8).ClearContents  ' Status
    Next r

    ws.Range("B34").ClearContents     ' Notes

    Application.ScreenUpdating = True
    MsgBox "Form cleared!", vbInformation, "QAA Quote System"
End Sub

Public Sub ClearExchange()
    Dim ws As Worksheet
    Dim r As Integer

    Set ws = ThisWorkbook.Sheets("Exchange")

    If MsgBox("Clear the Exchange quote form?" & vbCrLf & vbCrLf & _
              "This will reset all entries.", vbYesNo + vbQuestion, "QAA Quote System") = vbNo Then
        Exit Sub
    End If

    Application.ScreenUpdating = False

    ws.Range("B6").ClearContents      ' Customer email
    ws.Range("G6").Value = "Standard" ' Tier
    ws.Range("B7").Value = "No"       ' Approval
    ws.Range("D7").Value = "No"       ' Shipping
    ws.Range("F7").Value = "No"       ' Payment
    ws.Range("H7").Value = "No"       ' Core In Hand

    ' Clear line items
    For r = 11 To 22
        ws.Cells(r, 1).ClearContents  ' Part#
        ws.Cells(r, 4).ClearContents  ' Qty
    Next r

    ws.Range("A28").ClearContents     ' Notes

    Application.ScreenUpdating = True
    MsgBox "Form cleared!", vbInformation, "QAA Quote System"
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' UPDATE DATE MACRO (for revised quotes)
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub UpdateCustPropDate()
    ThisWorkbook.Sheets("Cust Property").Range("I5").Value = Date
    MsgBox "Updated date set to today!", vbInformation, "QAA Quote System"
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' HELPER FUNCTIONS
' ═══════════════════════════════════════════════════════════════════════════════

Private Function Count500HRUnits() As Integer
    Dim ws As Worksheet
    Dim cnt As Integer, r As Integer

    Set ws = ThisWorkbook.Sheets("500HR")
    cnt = 0

    For r = 10 To 15
        If Trim(ws.Cells(r, 2).Value) <> "" Then cnt = cnt + 1
    Next r

    Count500HRUnits = cnt
End Function

Private Function CountCPUnits() As Integer
    Dim ws As Worksheet
    Dim cnt As Integer, r As Integer

    Set ws = ThisWorkbook.Sheets("Cust Property")
    cnt = 0

    For r = 11 To 14
        If Trim(ws.Cells(r, 2).Value) <> "" Then cnt = cnt + 1
    Next r

    CountCPUnits = cnt
End Function

Private Function CountExItems() As Integer
    Dim ws As Worksheet
    Dim cnt As Integer, r As Integer

    Set ws = ThisWorkbook.Sheets("Exchange")
    cnt = 0

    For r = 11 To 22
        If Trim(ws.Cells(r, 1).Value) <> "" Then cnt = cnt + 1
    Next r

    CountExItems = cnt
End Function

' ═══════════════════════════════════════════════════════════════════════════════
' CHECKBOX TOGGLE (for findings)
' Call this from Worksheet_SelectionChange event
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub ToggleCheckbox(Target As Range)
    If Target.CountLarge > 1 Then Exit Sub

    Select Case Target.Value
        Case ChrW(9744)  ' ☐ unchecked
            Target.Value = ChrW(9745)  ' ☑ checked
        Case ChrW(9745)  ' ☑ checked
            Target.Value = ChrW(9744)  ' ☐ unchecked
    End Select
End Sub
'''

    return vba_code

def build_workbook(output_dir, logo_path):
    """Build the complete QAA Quote System workbook"""

    print("=" * 60)
    print("QAA QUOTE SYSTEM BUILDER")
    print("=" * 60)

    # Create button images
    print("\n[1/10] Creating button images...")
    create_button_images(output_dir)

    # Create workbook
    print("\n[2/10] Creating workbook...")
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create all sheets
    print("[3/10] Creating Menu sheet...")
    create_menu_sheet(wb, logo_path)

    print("[4/10] Creating 500HR sheet (using 5H suffix)...")
    create_500hr_sheet(wb)

    print("[5/10] Creating Cust Property sheet...")
    create_cust_property_sheet(wb)

    print("[6/10] Creating Exchange sheet...")
    create_exchange_sheet(wb)

    print("[7/10] Creating Parts Master...")
    create_parts_master_sheet(wb)

    print("[8/10] Creating supporting sheets...")
    create_bom_sheet(wb)
    create_config_sheet(wb)
    create_quote_log_sheet(wb)
    create_email_log_sheet(wb)
    create_inventory_sheet(wb)

    # Add button images to sheets
    print("[9/10] Adding button images to sheets...")

    btn_files = {
        'btn_generate_pdf': 'btn_generate_pdf.png',
        'btn_email': 'btn_email.png',
        'btn_clear': 'btn_clear.png',
        'btn_menu': 'btn_menu.png',
        'btn_500hr': 'btn_500hr.png',
        'btn_custprop': 'btn_custprop.png',
        'btn_exchange': 'btn_exchange.png',
        'btn_update': 'btn_update.png',
    }

    # Add buttons to Menu sheet
    ws_menu = wb['Menu']
    for btn_name, positions in [
        ('btn_500hr', 'B16'),
        ('btn_custprop', 'C16'),
        ('btn_exchange', 'D16'),
    ]:
        btn_path = os.path.join(output_dir, btn_files[btn_name])
        if os.path.exists(btn_path):
            try:
                img = XLImage(btn_path)
                ws_menu.add_image(img, positions)
            except Exception as e:
                print(f"  Warning: Could not add {btn_name}: {e}")

    # Add buttons to 500HR sheet
    ws_500hr = wb['500HR']
    for btn_name, pos in [('btn_generate_pdf', 'B38'), ('btn_email', 'D38'),
                          ('btn_clear', 'F38'), ('btn_menu', 'H38')]:
        btn_path = os.path.join(output_dir, btn_files[btn_name])
        if os.path.exists(btn_path):
            try:
                img = XLImage(btn_path)
                ws_500hr.add_image(img, pos)
            except Exception as e:
                print(f"  Warning: {e}")

    # Add buttons to Cust Property sheet
    ws_cp = wb['Cust Property']
    for btn_name, pos in [('btn_generate_pdf', 'B37'), ('btn_email', 'D37'),
                          ('btn_clear', 'F37'), ('btn_menu', 'H37'), ('btn_update', 'I5')]:
        btn_path = os.path.join(output_dir, btn_files.get(btn_name, ''))
        if btn_path and os.path.exists(btn_path):
            try:
                img = XLImage(btn_path)
                ws_cp.add_image(img, pos)
            except:
                pass

    # Add buttons to Exchange sheet
    ws_ex = wb['Exchange']
    for btn_name, pos in [('btn_generate_pdf', 'A31'), ('btn_email', 'C31'),
                          ('btn_clear', 'E31'), ('btn_menu', 'G31')]:
        btn_path = os.path.join(output_dir, btn_files[btn_name])
        if os.path.exists(btn_path):
            try:
                img = XLImage(btn_path)
                ws_ex.add_image(img, pos)
            except:
                pass

    # Add logo to sheets
    if logo_path and os.path.exists(logo_path):
        print("  Adding logo to sheets...")
        for sheet_name in ['500HR', 'Cust Property', 'Exchange']:
            try:
                logo = XLImage(logo_path)
                logo.width = 150
                logo.height = 50
                wb[sheet_name].add_image(logo, 'H1')
            except Exception as e:
                print(f"  Warning: Could not add logo to {sheet_name}: {e}")

    # Save workbook
    print("[10/10] Saving workbook...")
    xlsx_path = os.path.join(output_dir, 'QAA_Quote_System.xlsx')
    wb.save(xlsx_path)
    print(f"  Saved: {xlsx_path}")

    # Generate VBA module file
    vba_path = os.path.join(output_dir, 'QAA_VBA_MODULE.bas')
    with open(vba_path, 'w') as f:
        f.write(create_vba_module())
    print(f"  VBA Module: {vba_path}")

    print("\n" + "=" * 60)
    print("BUILD COMPLETE!")
    print("=" * 60)
    print(f"\nOutput files in: {output_dir}")
    print("\nNEXT STEPS:")
    print("1. Open QAA_Quote_System.xlsx in Excel")
    print("2. Save As > Excel Macro-Enabled Workbook (.xlsm)")
    print("3. Press Alt+F11 to open VBA Editor")
    print("4. Insert > Module")
    print("5. Copy/paste contents of QAA_VBA_MODULE.bas")
    print("6. Close VBA Editor")
    print("7. Right-click each button > Assign Macro")
    print("\nMACRO ASSIGNMENTS:")
    print("  - GENERATE PDF buttons → GenPDF500HR, GenPDFCustProp, GenPDFExchange")
    print("  - EMAIL buttons → Email500HR, EmailCustProp, EmailExchange")
    print("  - CLEAR buttons → Clear500HR, ClearCustProp, ClearExchange")
    print("  - MENU buttons → GoMenu")
    print("  - Navigation buttons → Go500HR, GoCustProp, GoExchange")
    print("  - UPDATE DATE → UpdateCustPropDate")

    return xlsx_path, vba_path

if __name__ == "__main__":
    output_dir = "./extracted/package"

    # Use the fixed PNG logo (converted from webp)
    logo_path = "./extracted/package/qaa_logo_header.png"
    if not os.path.exists(logo_path):
        logo_path = "./extracted/package/qaa_logo.png"

    build_workbook(output_dir, logo_path)
