#!/usr/bin/env python3
"""
QAA Quote System v13.0 - Professional Edition
- Professional PDF layout with logo, borders, footer
- Single intelligent Quote Tracker with follow-up workflow
- Reminder counting and weekly highlight system
- Status tracking (OPEN → CLOSED)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (Font, Alignment, Border, Side, PatternFill,
                              Protection, NamedStyle)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.comments import Comment
from PIL import Image, ImageDraw, ImageFont
from datetime import datetime

# =============================================================================
# PROFESSIONAL COLOR SCHEME
# =============================================================================
NAVY = "1B365D"          # Primary brand color
SKY_BLUE = "00A3E0"      # Accent color
GOLD = "C9A227"          # Highlight accent
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"    # Subtle backgrounds
MEDIUM_GRAY = "E0E0E0"   # Borders
DARK_GRAY = "666666"     # Secondary text
GREEN = "28A745"         # Success/Closed
RED = "DC3545"           # Alert/Needs attention
YELLOW = "FFF3CD"        # Warning highlight
INPUT_YELLOW = "FFFDE7"  # Input cells

# =============================================================================
# PROFESSIONAL STYLES
# =============================================================================
def create_styles():
    """Create reusable style definitions"""
    styles = {}

    # Header style (navy background, white text)
    styles['header'] = {
        'font': Font(name='Calibri', size=11, bold=True, color=WHITE),
        'fill': PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin', color=NAVY),
            right=Side(style='thin', color=NAVY),
            top=Side(style='thin', color=NAVY),
            bottom=Side(style='thin', color=NAVY)
        )
    }

    # Subheader style
    styles['subheader'] = {
        'font': Font(name='Calibri', size=10, bold=True, color=NAVY),
        'fill': PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': Border(
            left=Side(style='thin', color=MEDIUM_GRAY),
            right=Side(style='thin', color=MEDIUM_GRAY),
            top=Side(style='thin', color=MEDIUM_GRAY),
            bottom=Side(style='thin', color=MEDIUM_GRAY)
        )
    }

    # Input cell style
    styles['input'] = {
        'font': Font(name='Calibri', size=10),
        'fill': PatternFill(start_color=INPUT_YELLOW, end_color=INPUT_YELLOW, fill_type='solid'),
        'border': Border(
            left=Side(style='thin', color=MEDIUM_GRAY),
            right=Side(style='thin', color=MEDIUM_GRAY),
            top=Side(style='thin', color=MEDIUM_GRAY),
            bottom=Side(style='thin', color=MEDIUM_GRAY)
        )
    }

    # Data cell style
    styles['data'] = {
        'font': Font(name='Calibri', size=10),
        'alignment': Alignment(vertical='center'),
        'border': Border(
            left=Side(style='thin', color=MEDIUM_GRAY),
            right=Side(style='thin', color=MEDIUM_GRAY),
            top=Side(style='thin', color=MEDIUM_GRAY),
            bottom=Side(style='thin', color=MEDIUM_GRAY)
        )
    }

    # Currency style
    styles['currency'] = {
        'font': Font(name='Calibri', size=10),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'number_format': '$#,##0.00',
        'border': Border(
            left=Side(style='thin', color=MEDIUM_GRAY),
            right=Side(style='thin', color=MEDIUM_GRAY),
            top=Side(style='thin', color=MEDIUM_GRAY),
            bottom=Side(style='thin', color=MEDIUM_GRAY)
        )
    }

    # Total style (highlighted)
    styles['total'] = {
        'font': Font(name='Calibri', size=12, bold=True, color=NAVY),
        'fill': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
        'alignment': Alignment(horizontal='right', vertical='center'),
        'number_format': '$#,##0.00',
        'border': Border(
            left=Side(style='medium', color=NAVY),
            right=Side(style='medium', color=NAVY),
            top=Side(style='medium', color=NAVY),
            bottom=Side(style='medium', color=NAVY)
        )
    }

    return styles

STYLES = create_styles()

def apply_style(cell, style_name):
    """Apply a predefined style to a cell"""
    style = STYLES.get(style_name, STYLES['data'])
    for attr, value in style.items():
        if attr == 'number_format':
            cell.number_format = value
        else:
            setattr(cell, attr, value)

# =============================================================================
# BUTTON IMAGE CREATION
# =============================================================================
def create_professional_buttons(output_dir):
    """Create sleek, professional button images"""

    buttons = {
        'btn_generate_pdf': ('GENERATE PDF', (160, 40), NAVY),
        'btn_email': ('SEND EMAIL', (140, 40), NAVY),
        'btn_clear': ('CLEAR FORM', (130, 40), DARK_GRAY),
        'btn_menu': ('← MENU', (100, 36), SKY_BLUE),
        'btn_update': ('UPDATE', (100, 36), NAVY),
        'btn_500hr': ('500HR SERVICE', (150, 44), NAVY),
        'btn_custprop': ('CUST PROPERTY', (150, 44), NAVY),
        'btn_exchange': ('EXCHANGE/SALE', (150, 44), NAVY),
        'btn_add_reminder': ('+ REMINDER', (120, 36), 'E67E22'),
        'btn_close_quote': ('CLOSE QUOTE', (120, 36), GREEN),
    }

    for name, (text, size, color) in buttons.items():
        img = Image.new('RGBA', size, (0, 0, 0, 0))
        draw = ImageDraw.Draw(img)

        # Convert hex color to RGB
        if isinstance(color, str):
            color_rgb = tuple(int(color[i:i+2], 16) for i in (0, 2, 4))
        else:
            color_rgb = color

        # Draw rounded rectangle with subtle gradient effect
        radius = 6
        draw.rounded_rectangle(
            [(0, 0), (size[0]-1, size[1]-1)],
            radius=radius,
            fill=color_rgb,
            outline=tuple(max(0, c-30) for c in color_rgb),
            width=1
        )

        # Add text
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 12)
        except:
            font = ImageFont.load_default()

        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        x = (size[0] - text_width) // 2
        y = (size[1] - text_height) // 2 - 1

        draw.text((x, y), text, fill=(255, 255, 255), font=font)

        filepath = os.path.join(output_dir, f'{name}.png')
        img.save(filepath, 'PNG')
        print(f"  Created: {name}.png")

    return list(buttons.keys())

# =============================================================================
# PROFESSIONAL PDF-READY 500HR SHEET
# =============================================================================
def create_500hr_sheet(wb):
    """Create professional 500HR quote sheet optimized for PDF output"""
    ws = wb.create_sheet("500HR")

    # Column widths for professional layout
    widths = {'A': 3, 'B': 16, 'C': 12, 'D': 28, 'E': 11, 'F': 9, 'G': 14, 'H': 11, 'I': 12, 'J': 3}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Row heights
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 15

    # =========================================================================
    # HEADER SECTION (Professional letterhead style)
    # =========================================================================

    # Company name
    ws.merge_cells('B1:I1')
    ws['B1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B1'].font = Font(name='Calibri', size=18, bold=True, color=NAVY)
    ws['B1'].alignment = Alignment(horizontal='center', vertical='center')

    # Tagline
    ws.merge_cells('B2:I2')
    ws['B2'] = "A Hartzell Engine Technologies Company"
    ws['B2'].font = Font(name='Calibri', size=10, italic=True, color=SKY_BLUE)
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

    # Document title bar
    ws.merge_cells('B3:I3')
    ws['B3'] = "500HR MAGNETO INSPECTION SERVICE QUOTE"
    apply_style(ws['B3'], 'header')
    ws['B3'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    ws.row_dimensions[3].height = 28

    # =========================================================================
    # QUOTE INFO SECTION
    # =========================================================================

    # Quote number and date row
    ws['B5'] = "Quote #:"
    ws['B5'].font = Font(bold=True, size=10, color=NAVY)
    ws['C5'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['C5'].font = Font(bold=True, size=11, color=NAVY)

    ws['F5'] = "Date:"
    ws['F5'].font = Font(bold=True, size=10, color=NAVY)
    ws['G5'] = datetime.now().strftime("%m/%d/%Y")
    ws['G5'].number_format = 'MM/DD/YYYY'

    ws['H5'] = "Valid:"
    ws['H5'].font = Font(bold=True, size=10, color=NAVY)
    ws['I5'] = "30 Days"

    # Customer email row
    ws['B6'] = "Customer Email:"
    ws['B6'].font = Font(bold=True, size=10, color=NAVY)
    ws.merge_cells('C6:E6')
    ws['C6'] = ""
    apply_style(ws['C6'], 'input')

    ws['G6'] = "Approval Req?"
    ws['G6'].font = Font(bold=True, size=9, color=NAVY)
    ws['H6'] = "No"
    apply_style(ws['H6'], 'input')

    # Add Yes/No validation
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv_yesno)
    dv_yesno.add(ws['H6'])

    # Horizontal line
    for col in range(2, 10):
        ws.cell(row=7, column=col).border = Border(bottom=Side(style='thin', color=NAVY))
    ws.row_dimensions[7].height = 8

    # =========================================================================
    # MAGNETO TABLE (Professional styling)
    # =========================================================================

    # Table header
    headers = [('B', 'Magneto P/N'), ('C', 'Serial #'), ('D', 'Description'),
               ('E', '500HR Price'), ('F', 'Ovhl?'), ('G', 'CO Part #'),
               ('H', 'CO Price'), ('I', 'Line Total')]

    for col, header in headers:
        cell = ws[f'{col}8']
        cell.value = header
        apply_style(cell, 'header')
    ws.row_dimensions[8].height = 24

    # Magneto data rows (6 units)
    for row in range(9, 15):
        # P/N input
        ws[f'B{row}'] = ""
        apply_style(ws[f'B{row}'], 'input')

        # Serial input
        ws[f'C{row}'] = ""
        apply_style(ws[f'C{row}'], 'input')

        # Description lookup - USING 5H SUFFIX (NOT -5H!)
        ws[f'D{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row}&"5H",\'Parts Master\'!$A:$H,2,FALSE),IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found")))'
        apply_style(ws[f'D{row}'], 'data')

        # 500HR Price - USING 5H SUFFIX
        ws[f'E{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row}&"5H",\'Parts Master\'!$A:$H,4,FALSE),Config!$B$5))'
        apply_style(ws[f'E{row}'], 'currency')

        # Overhaul Required
        ws[f'F{row}'] = ""
        apply_style(ws[f'F{row}'], 'input')
        ws[f'F{row}'].alignment = Alignment(horizontal='center')
        dv_yesno.add(ws[f'F{row}'])

        # CO Part Number
        ws[f'G{row}'] = f'=IF(F{row}="Yes",SUBSTITUTE(B{row},"5H","")&"-CO","")'
        apply_style(ws[f'G{row}'], 'data')

        # CO Price
        ws[f'H{row}'] = f'=IF(G{row}="","",IFERROR(VLOOKUP(G{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        apply_style(ws[f'H{row}'], 'currency')

        # Line Total
        ws[f'I{row}'] = f'=IF(B{row}="","",IF(F{row}="Yes",H{row},E{row}))'
        apply_style(ws[f'I{row}'], 'currency')

        # Alternate row shading
        if row % 2 == 0:
            for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                if col not in ['B', 'C', 'F']:  # Keep input cells yellow
                    ws[f'{col}{row}'].fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')

    # =========================================================================
    # FINDINGS SECTION
    # =========================================================================

    ws.merge_cells('B16:I16')
    ws['B16'] = "INSPECTION FINDINGS"
    ws['B16'].font = Font(bold=True, size=11, color=WHITE)
    ws['B16'].fill = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type='solid')
    ws['B16'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[16].height = 22

    # Findings headers
    findings = ["Coil", "Capacitor", "Dist Block", "Impulse", "Gear", "Points", "Bearings", "Oil ⚠"]
    for i, finding in enumerate(findings):
        col = get_column_letter(2 + i)
        ws[f'{col}17'] = finding
        ws[f'{col}17'].font = Font(size=8, bold=True, color=NAVY)
        ws[f'{col}17'].alignment = Alignment(horizontal='center')
        ws[f'{col}17'].border = Border(bottom=Side(style='thin', color=MEDIUM_GRAY))

    # Checkbox grid
    for row in range(18, 24):
        for col_num in range(2, 10):
            col = get_column_letter(col_num)
            ws[f'{col}{row}'] = "☐"
            ws[f'{col}{row}'].font = Font(size=12)
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
            ws[f'{col}{row}'].border = Border(
                left=Side(style='hair', color=MEDIUM_GRAY),
                right=Side(style='hair', color=MEDIUM_GRAY),
                top=Side(style='hair', color=MEDIUM_GRAY),
                bottom=Side(style='hair', color=MEDIUM_GRAY)
            )

    # =========================================================================
    # TOTALS SECTION
    # =========================================================================

    # Findings add-on
    ws['B25'] = "Findings Add-On:"
    ws['B25'].font = Font(size=9, color=DARK_GRAY)
    ws['D25'] = '=SUMPRODUCT((B18:I23<>"☐")*1)/8*Config!$B$6'
    ws['D25'].number_format = '$#,##0.00'
    ws['D25'].font = Font(size=9)

    # Horizontal line before totals
    for col in range(2, 10):
        ws.cell(row=26, column=col).border = Border(bottom=Side(style='thin', color=NAVY))
    ws.row_dimensions[26].height = 6

    # Service total
    ws['F27'] = "500HR Service:"
    ws['F27'].font = Font(bold=True, size=10, color=NAVY)
    ws['F27'].alignment = Alignment(horizontal='right')
    ws['I27'] = '=SUMIF(F9:F14,"<>Yes",E9:E14)'
    apply_style(ws['I27'], 'currency')

    # Overhaul total
    ws['F28'] = "Overhaul Total:"
    ws['F28'].font = Font(bold=True, size=10, color=NAVY)
    ws['F28'].alignment = Alignment(horizontal='right')
    ws['I28'] = '=SUMIF(F9:F14,"Yes",H9:H14)'
    apply_style(ws['I28'], 'currency')

    # Grand total
    ws.merge_cells('F29:G29')
    ws['F29'] = "GRAND TOTAL:"
    ws['F29'].font = Font(bold=True, size=12, color=NAVY)
    ws['F29'].alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('H29:I29')
    ws['H29'] = '=I27+I28+D25'
    apply_style(ws['H29'], 'total')
    ws.row_dimensions[29].height = 28

    # =========================================================================
    # NOTES SECTION
    # =========================================================================

    ws['B31'] = "Notes / Special Instructions:"
    ws['B31'].font = Font(bold=True, size=9, color=NAVY)

    ws.merge_cells('B32:I33')
    ws['B32'] = ""
    apply_style(ws['B32'], 'input')
    ws['B32'].alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[32].height = 40

    # =========================================================================
    # FOOTER (Professional)
    # =========================================================================

    ws.merge_cells('B35:I35')
    ws['B35'] = "Thank you for choosing Quality Aircraft Accessories!"
    ws['B35'].font = Font(italic=True, size=10, color=NAVY)
    ws['B35'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B36:I36')
    ws['B36'] = "5746 E Apache St, Tulsa, OK 74115 | 918-835-6948 | sales@qaainc.com | qaa.com"
    ws['B36'].font = Font(size=9, color=DARK_GRAY)
    ws['B36'].alignment = Alignment(horizontal='center')

    # Terms footer
    ws.merge_cells('B37:I37')
    ws['B37'] = "Quote valid for 30 days. Prices subject to change. Core charges may apply for exchange units."
    ws['B37'].font = Font(size=8, italic=True, color=DARK_GRAY)
    ws['B37'].alignment = Alignment(horizontal='center')

    # Print area for PDF
    ws.print_area = 'A1:J37'
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    return ws

# =============================================================================
# CUSTOMER PROPERTY SHEET
# =============================================================================
def create_cust_property_sheet(wb):
    """Create Customer Property Overhaul sheet"""
    ws = wb.create_sheet("Cust Property")

    widths = {'A': 8, 'B': 16, 'C': 24, 'D': 10, 'E': 11, 'F': 11, 'G': 11, 'H': 9, 'I': 11, 'J': 3}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header
    ws.merge_cells('B1:I1')
    ws['B1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B1'].font = Font(name='Calibri', size=18, bold=True, color=NAVY)
    ws['B1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B2:I2')
    ws['B2'] = "A Hartzell Engine Technologies Company"
    ws['B2'].font = Font(name='Calibri', size=10, italic=True, color=SKY_BLUE)
    ws['B2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B3:I3')
    ws['B3'] = "CUSTOMER PROPERTY OVERHAUL QUOTE"
    apply_style(ws['B3'], 'header')
    ws['B3'].font = Font(name='Calibri', size=12, bold=True, color=WHITE)
    ws.row_dimensions[3].height = 28

    # Quote info
    ws['B5'] = "Quote #:"
    ws['B5'].font = Font(bold=True, color=NAVY)
    ws['C5'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['C5'].font = Font(bold=True, size=11, color=NAVY)

    ws['E5'] = "Date:"
    ws['E5'].font = Font(bold=True, color=NAVY)
    ws['F5'] = datetime.now().strftime("%m/%d/%Y")

    ws['G5'] = "Updated:"
    ws['G5'].font = Font(bold=True, color=NAVY)
    ws['H5'] = ""
    apply_style(ws['H5'], 'input')

    # Customer email
    ws['B6'] = "Customer Email:"
    ws['B6'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('C6:E6')
    ws['C6'] = ""
    apply_style(ws['C6'], 'input')

    ws['F6'] = "Tier:"
    ws['F6'].font = Font(bold=True, color=NAVY)
    ws['G6'] = "Standard"
    apply_style(ws['G6'], 'input')

    dv_tier = DataValidation(type="list", formula1='"Standard,Tier 1,Tier 2,Tier 3"')
    ws.add_data_validation(dv_tier)
    dv_tier.add(ws['G6'])

    # Requirements row
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(dv_yesno)

    ws['B7'] = "Needs:"
    ws['B7'].font = Font(bold=True, size=9, color=NAVY)
    ws['C7'] = "Approval?"
    ws['D7'] = "No"
    apply_style(ws['D7'], 'input')
    dv_yesno.add(ws['D7'])

    ws['E7'] = "Shipping?"
    ws['F7'] = "No"
    apply_style(ws['F7'], 'input')
    dv_yesno.add(ws['F7'])

    ws['G7'] = "Payment?"
    ws['H7'] = "No"
    apply_style(ws['H7'], 'input')
    dv_yesno.add(ws['H7'])

    # Separator
    for col in range(2, 10):
        ws.cell(row=8, column=col).border = Border(bottom=Side(style='thin', color=NAVY))
    ws.row_dimensions[8].height = 6

    # Units table header
    headers = [('B', 'Unit P/N'), ('C', 'Description'), ('D', 'Type'),
               ('E', 'Base Price'), ('F', 'Your Price'), ('G', 'Exch Cap')]
    for col, header in headers:
        cell = ws[f'{col}9']
        cell.value = header
        apply_style(cell, 'header')
    ws.row_dimensions[9].height = 24

    # Unit rows (4 units)
    for row in range(10, 14):
        ws[f'B{row}'] = ""
        apply_style(ws[f'B{row}'], 'input')

        ws[f'C{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found"))'
        apply_style(ws[f'C{row}'], 'data')

        ws[f'D{row}'] = f'=IF(B{row}="","",IF(COUNTIF(BOM!$A:$A,SUBSTITUTE(B{row},"-CO",""))>0,"In-House","Vendor"))'
        apply_style(ws[f'D{row}'], 'data')
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        ws[f'E{row}'] = ""
        apply_style(ws[f'E{row}'], 'input')
        ws[f'E{row}'].number_format = '$#,##0.00'

        ws[f'F{row}'] = f'=IF(E{row}="","",E{row}*(1-VLOOKUP($G$6,Config!$A$10:$B$13,2,FALSE)))'
        apply_style(ws[f'F{row}'], 'currency')

        ws[f'G{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(SUBSTITUTE(B{row},"-CO","-OH"),\'Parts Master\'!$A:$H,4,FALSE),"N/A"))'
        apply_style(ws[f'G{row}'], 'currency')

    # O&A Parts section
    ws.merge_cells('B15:I15')
    ws['B15'] = "OVER & ABOVE PARTS"
    ws['B15'].font = Font(bold=True, size=11, color=WHITE)
    ws['B15'].fill = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type='solid')
    ws['B15'].alignment = Alignment(horizontal='center')

    # O&A headers
    oa_headers = [('A', '#'), ('B', 'Part #'), ('C', 'Description'), ('D', 'Qty'),
                  ('E', 'Unit $'), ('F', 'Override'), ('G', 'Extended'), ('H', 'Status')]
    for col, header in oa_headers:
        cell = ws[f'{col}16']
        cell.value = header
        apply_style(cell, 'subheader')

    # O&A rows (10 parts)
    for row in range(17, 27):
        ws[f'A{row}'] = ""
        apply_style(ws[f'A{row}'], 'input')

        ws[f'B{row}'] = ""
        apply_style(ws[f'B{row}'], 'input')

        ws[f'C{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,2,FALSE),""))'
        apply_style(ws[f'C{row}'], 'data')

        ws[f'D{row}'] = ""
        apply_style(ws[f'D{row}'], 'input')
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        ws[f'E{row}'] = f'=IF(B{row}="","",IFERROR(VLOOKUP(B{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        apply_style(ws[f'E{row}'], 'currency')

        ws[f'F{row}'] = ""
        apply_style(ws[f'F{row}'], 'input')
        ws[f'F{row}'].number_format = '$#,##0.00'

        ws[f'G{row}'] = f'=IF(D{row}="","",IF(F{row}<>"",F{row}*D{row},E{row}*D{row}))'
        apply_style(ws[f'G{row}'], 'currency')

        ws[f'H{row}'] = ""
        apply_style(ws[f'H{row}'], 'input')

    # Totals
    ws['E28'] = "Labor Total:"
    ws['E28'].font = Font(bold=True, color=NAVY)
    ws['E28'].alignment = Alignment(horizontal='right')
    ws['G28'] = '=SUM(F10:F13)'
    apply_style(ws['G28'], 'currency')

    ws['E29'] = "Parts Total:"
    ws['E29'].font = Font(bold=True, color=NAVY)
    ws['E29'].alignment = Alignment(horizontal='right')
    ws['G29'] = '=SUM(G17:G26)'
    apply_style(ws['G29'], 'currency')

    ws.merge_cells('E30:F30')
    ws['E30'] = "GRAND TOTAL:"
    ws['E30'].font = Font(bold=True, size=12, color=NAVY)
    ws['E30'].alignment = Alignment(horizontal='right')
    ws['G30'] = '=G28+G29'
    apply_style(ws['G30'], 'total')

    # Notes
    ws['B32'] = "Notes:"
    ws['B32'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('B33:H34')
    ws['B33'] = ""
    apply_style(ws['B33'], 'input')
    ws['B33'].alignment = Alignment(vertical='top', wrap_text=True)

    # Footer
    ws.merge_cells('B36:H36')
    ws['B36'] = "5746 E Apache St, Tulsa, OK 74115 | 918-835-6948 | sales@qaainc.com | qaa.com"
    ws['B36'].font = Font(size=9, color=DARK_GRAY)
    ws['B36'].alignment = Alignment(horizontal='center')

    ws.print_area = 'A1:I36'

    return ws

# =============================================================================
# EXCHANGE SHEET
# =============================================================================
def create_exchange_sheet(wb):
    """Create Exchange/Inventory Sale sheet"""
    ws = wb.create_sheet("Exchange")

    widths = {'A': 16, 'B': 24, 'C': 6, 'D': 5, 'E': 11, 'F': 11, 'G': 11, 'H': 11, 'I': 8, 'J': 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header
    ws.merge_cells('A1:J1')
    ws['A1'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color=NAVY)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = "A Hartzell Engine Technologies Company"
    ws['A2'].font = Font(size=10, italic=True, color=SKY_BLUE)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:J3')
    ws['A3'] = "EXCHANGE / INVENTORY SALE QUOTE"
    apply_style(ws['A3'], 'header')
    ws['A3'].font = Font(size=12, bold=True, color=WHITE)
    ws.row_dimensions[3].height = 28

    # Quote info
    ws['A5'] = "Quote #:"
    ws['A5'].font = Font(bold=True, color=NAVY)
    ws['B5'] = '=Config!B2&"-"&TEXT(Config!B3,"00000")'
    ws['B5'].font = Font(bold=True, size=11, color=NAVY)

    ws['D5'] = "Date:"
    ws['D5'].font = Font(bold=True, color=NAVY)
    ws['E5'] = datetime.now().strftime("%m/%d/%Y")

    ws['G5'] = "Valid:"
    ws['G5'].font = Font(bold=True, color=NAVY)
    ws['H5'] = "30 Days"

    # Customer email
    ws['A6'] = "Customer Email:"
    ws['A6'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('B6:D6')
    ws['B6'] = ""
    apply_style(ws['B6'], 'input')

    ws['F6'] = "Tier:"
    ws['F6'].font = Font(bold=True, color=NAVY)
    ws['G6'] = "Standard"
    apply_style(ws['G6'], 'input')

    dv_tier = DataValidation(type="list", formula1='"Standard,Tier 1,Tier 2,Tier 3"')
    ws.add_data_validation(dv_tier)
    dv_tier.add(ws['G6'])

    # Requirements
    dv_yesno = DataValidation(type="list", formula1='"Yes,No"')
    ws.add_data_validation(dv_yesno)

    ws['A7'] = "Approval?"
    ws['B7'] = "No"
    apply_style(ws['B7'], 'input')
    dv_yesno.add(ws['B7'])

    ws['C7'] = "Ship?"
    ws['D7'] = "No"
    apply_style(ws['D7'], 'input')
    dv_yesno.add(ws['D7'])

    ws['E7'] = "Payment?"
    ws['F7'] = "No"
    apply_style(ws['F7'], 'input')
    dv_yesno.add(ws['F7'])

    ws['G7'] = "Core In Hand?"
    ws['H7'] = "No"
    apply_style(ws['H7'], 'input')
    dv_yesno.add(ws['H7'])

    # Separator
    for col in range(1, 11):
        ws.cell(row=8, column=col).border = Border(bottom=Side(style='thin', color=NAVY))

    # Line items header
    headers = [('A', 'Part #'), ('B', 'Description'), ('C', 'Cond'), ('D', 'Qty'),
               ('E', 'List $'), ('F', 'Your $'), ('G', 'Extended'), ('H', 'Core $'),
               ('I', 'Stock'), ('J', 'Lead')]
    for col, header in headers:
        cell = ws[f'{col}9']
        cell.value = header
        apply_style(cell, 'header')
    ws.row_dimensions[9].height = 24

    # Line item rows (12 items)
    for row in range(10, 22):
        ws[f'A{row}'] = ""
        apply_style(ws[f'A{row}'], 'input')

        ws[f'B{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,2,FALSE),"Not Found"))'
        apply_style(ws[f'B{row}'], 'data')

        ws[f'C{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,3,FALSE),""))'
        apply_style(ws[f'C{row}'], 'data')
        ws[f'C{row}'].alignment = Alignment(horizontal='center')

        ws[f'D{row}'] = ""
        apply_style(ws[f'D{row}'], 'input')
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        ws[f'E{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,4,FALSE),0))'
        apply_style(ws[f'E{row}'], 'currency')

        ws[f'F{row}'] = f'=IF(E{row}="","",E{row}*(1-VLOOKUP($G$6,Config!$A$10:$B$13,2,FALSE)))'
        apply_style(ws[f'F{row}'], 'currency')

        ws[f'G{row}'] = f'=IF(D{row}="","",F{row}*D{row})'
        apply_style(ws[f'G{row}'], 'currency')

        ws[f'H{row}'] = f'=IF(A{row}="","",IF($H$7="Yes",0,IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,5,FALSE),0)*D{row}))'
        apply_style(ws[f'H{row}'], 'currency')

        ws[f'I{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,6,FALSE),""))'
        apply_style(ws[f'I{row}'], 'data')
        ws[f'I{row}'].alignment = Alignment(horizontal='center')

        ws[f'J{row}'] = f'=IF(A{row}="","",IFERROR(VLOOKUP(A{row},\'Parts Master\'!$A:$H,7,FALSE),""))'
        apply_style(ws[f'J{row}'], 'data')

    # Totals section
    ws['A23'] = "Subtotal:"
    ws['A23'].font = Font(bold=True, color=NAVY)
    ws['B23'] = '=SUM(G10:G21)'
    apply_style(ws['B23'], 'currency')

    ws['D23'] = "Core:"
    ws['D23'].font = Font(bold=True, color=NAVY)
    ws['E23'] = '=SUM(H10:H21)'
    apply_style(ws['E23'], 'currency')

    ws['F23'] = '=IF(H7="Yes","(WAIVED)","")'
    ws['F23'].font = Font(size=9, color=GREEN)

    ws['G23'] = "Savings:"
    ws['G23'].font = Font(bold=True, color=GREEN)
    ws['H23'] = '=SUM(E10:E21)-SUM(F10:F21)'
    ws['H23'].font = Font(bold=True, color=GREEN)
    ws['H23'].number_format = '$#,##0.00'

    # Grand total
    ws.merge_cells('D24:E24')
    ws['D24'] = "GRAND TOTAL:"
    ws['D24'].font = Font(bold=True, size=12, color=NAVY)
    ws['D24'].alignment = Alignment(horizontal='right')

    ws.merge_cells('F24:G24')
    ws['F24'] = '=B23+IF(H7="Yes",0,E23)'
    apply_style(ws['F24'], 'total')

    # Notes
    ws['A26'] = "Notes:"
    ws['A26'].font = Font(bold=True, color=NAVY)
    ws.merge_cells('A27:J28')
    ws['A27'] = ""
    apply_style(ws['A27'], 'input')

    # Footer
    ws.merge_cells('A30:J30')
    ws['A30'] = "Core charges refunded upon receipt of serviceable core within 30 days."
    ws['A30'].font = Font(size=9, italic=True, color=DARK_GRAY)
    ws['A30'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A31:J31')
    ws['A31'] = "5746 E Apache St, Tulsa, OK 74115 | 918-835-6948 | sales@qaainc.com | qaa.com"
    ws['A31'].font = Font(size=9, color=DARK_GRAY)
    ws['A31'].alignment = Alignment(horizontal='center')

    ws.print_area = 'A1:J31'
    ws.page_setup.orientation = 'landscape'

    return ws

# =============================================================================
# INTELLIGENT QUOTE TRACKER (Single Log with Follow-up Workflow)
# =============================================================================
def create_quote_tracker_sheet(wb):
    """
    Create intelligent Quote Tracker with:
    - Contact tracking
    - Follow-up reminders
    - Status workflow (OPEN → CLOSED)
    - Weekly highlight for overdue
    - Reminder count
    """
    ws = wb.create_sheet("Quote Tracker")

    # Column widths
    widths = {
        'A': 14,   # Quote #
        'B': 11,   # Date Created
        'C': 25,   # Customer Email
        'D': 12,   # Quote Type
        'E': 12,   # Total Amount
        'F': 10,   # Approval
        'G': 10,   # Shipping
        'H': 10,   # Payment
        'I': 10,   # Reminders Sent
        'J': 12,   # Last Contact
        'K': 10,   # Days Since
        'L': 10,   # Status
        'M': 40,   # Notes/Email Ref
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = "QUOTE TRACKER - Follow-up & Status Management"
    ws['A1'].font = Font(size=14, bold=True, color=WHITE)
    ws['A1'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Instructions row
    ws.merge_cells('A2:M2')
    ws['A2'] = "★ RED rows = needs follow-up (7+ days since last contact) | ★ Mark needs as DONE when fulfilled | ★ Set Status to CLOSED when complete"
    ws['A2'].font = Font(size=9, italic=True, color=DARK_GRAY)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        ('A', 'Quote #'),
        ('B', 'Created'),
        ('C', 'Customer Email'),
        ('D', 'Type'),
        ('E', 'Amount'),
        ('F', 'Approval'),
        ('G', 'Shipping'),
        ('H', 'Payment'),
        ('I', 'Reminders'),
        ('J', 'Last Contact'),
        ('K', 'Days Since'),
        ('L', 'Status'),
        ('M', 'Notes / Email Reference'),
    ]

    for col, header in headers:
        cell = ws[f'{col}3']
        cell.value = header
        apply_style(cell, 'header')
    ws.row_dimensions[3].height = 28

    # Data validation for needs columns
    dv_needs = DataValidation(type="list", formula1='"Yes,No,DONE"', allow_blank=True)
    ws.add_data_validation(dv_needs)

    dv_status = DataValidation(type="list", formula1='"OPEN,CLOSED,ON HOLD"', allow_blank=True)
    ws.add_data_validation(dv_status)

    # Pre-format data rows (50 rows)
    thin_border = Border(
        left=Side(style='thin', color=MEDIUM_GRAY),
        right=Side(style='thin', color=MEDIUM_GRAY),
        top=Side(style='thin', color=MEDIUM_GRAY),
        bottom=Side(style='thin', color=MEDIUM_GRAY)
    )

    for row in range(4, 54):
        # Apply border to all columns
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            ws[f'{col}{row}'].border = thin_border

        # Created date
        ws[f'B{row}'].number_format = 'MM/DD/YY'

        # Type
        ws[f'D{row}'].alignment = Alignment(horizontal='center')

        # Amount
        ws[f'E{row}'].number_format = '$#,##0.00'

        # Approval, Shipping, Payment (needs columns)
        for col in ['F', 'G', 'H']:
            ws[f'{col}{row}'].alignment = Alignment(horizontal='center')
            dv_needs.add(ws[f'{col}{row}'])

        # Reminders count
        ws[f'I{row}'].alignment = Alignment(horizontal='center')

        # Last Contact date
        ws[f'J{row}'].number_format = 'MM/DD/YY'

        # Days Since (auto-calculated)
        ws[f'K{row}'] = f'=IF(J{row}="","",TODAY()-J{row})'
        ws[f'K{row}'].alignment = Alignment(horizontal='center')

        # Status
        ws[f'L{row}'].alignment = Alignment(horizontal='center')
        dv_status.add(ws[f'L{row}'])

        # Notes
        ws[f'M{row}'].alignment = Alignment(wrap_text=True)

    # Conditional formatting: Highlight rows needing follow-up (7+ days, still OPEN)
    red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

    # Red: OPEN and 7+ days since contact
    ws.conditional_formatting.add(
        'A4:M53',
        FormulaRule(formula=['AND($L4="OPEN",$K4>=7)'], fill=red_fill)
    )

    # Yellow: OPEN and 3-6 days
    ws.conditional_formatting.add(
        'A4:M53',
        FormulaRule(formula=['AND($L4="OPEN",$K4>=3,$K4<7)'], fill=yellow_fill)
    )

    # Green: CLOSED
    ws.conditional_formatting.add(
        'A4:M53',
        FormulaRule(formula=['$L4="CLOSED"'], fill=green_fill)
    )

    # Summary section at top
    ws['A55'] = "SUMMARY"
    ws['A55'].font = Font(bold=True, color=NAVY)

    ws['A56'] = "Open Quotes:"
    ws['B56'] = '=COUNTIF(L4:L53,"OPEN")'
    ws['B56'].font = Font(bold=True)

    ws['C56'] = "Need Follow-up:"
    ws['D56'] = '=SUMPRODUCT((L4:L53="OPEN")*(K4:K53>=7))'
    ws['D56'].font = Font(bold=True, color=RED)

    ws['E56'] = "Closed This Month:"
    ws['F56'] = '=SUMPRODUCT((L4:L53="CLOSED")*(MONTH(B4:B53)=MONTH(TODAY())))'

    # Freeze panes
    ws.freeze_panes = 'A4'

    # Auto-filter
    ws.auto_filter.ref = 'A3:M53'

    return ws

# =============================================================================
# MENU SHEET
# =============================================================================
def create_menu_sheet(wb, logo_path):
    """Create navigation menu"""
    ws = wb.create_sheet("Menu", 0)

    widths = {'A': 5, 'B': 22, 'C': 22, 'D': 22, 'E': 22, 'F': 5}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Header
    ws.merge_cells('B2:E2')
    ws['B2'] = "QUALITY AIRCRAFT ACCESSORIES"
    ws['B2'].font = Font(size=22, bold=True, color=NAVY)
    ws['B2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('B3:E3')
    ws['B3'] = "A Hartzell Engine Technologies Company"
    ws['B3'].font = Font(size=11, italic=True, color=SKY_BLUE)
    ws['B3'].alignment = Alignment(horizontal='center')

    # Logo
    if logo_path and os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            img.width = 180
            img.height = 70
            ws.add_image(img, 'B5')
        except Exception as e:
            print(f"  Logo warning: {e}")

    # System title
    ws.merge_cells('B9:E9')
    ws['B9'] = "QUOTE SYSTEM v13.0"
    ws['B9'].font = Font(size=18, bold=True, color=WHITE)
    ws['B9'].fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type='solid')
    ws['B9'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[9].height = 40

    # Quote type cards
    ws['B11'] = "SELECT QUOTE TYPE:"
    ws['B11'].font = Font(size=12, bold=True, color=NAVY)

    # Card headers
    for col, text in [('B', '500HR SERVICE'), ('C', 'CUST PROPERTY'), ('D', 'EXCHANGE/SALE')]:
        ws[f'{col}13'] = text
        ws[f'{col}13'].font = Font(bold=True, size=11, color=NAVY)
        ws[f'{col}13'].alignment = Alignment(horizontal='center')

    # Card descriptions
    for col, text in [('B', 'Magneto inspection quotes'),
                      ('C', 'Customer unit overhauls'),
                      ('D', 'Exchange & inventory')]:
        ws[f'{col}14'] = text
        ws[f'{col}14'].font = Font(size=9, color=DARK_GRAY)
        ws[f'{col}14'].alignment = Alignment(horizontal='center')

    # Button placeholders
    for col in ['B', 'C', 'D']:
        ws[f'{col}16'] = "[Button]"
        ws[f'{col}16'].alignment = Alignment(horizontal='center')

    # Quote Tracker link
    ws.merge_cells('B19:E19')
    ws['B19'] = "QUOTE TRACKER"
    ws['B19'].font = Font(size=14, bold=True, color=WHITE)
    ws['B19'].fill = PatternFill(start_color=SKY_BLUE, end_color=SKY_BLUE, fill_type='solid')
    ws['B19'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[19].height = 35

    ws.merge_cells('B20:E20')
    ws['B20'] = "Track follow-ups, reminders, and quote status"
    ws['B20'].font = Font(size=10, color=DARK_GRAY)
    ws['B20'].alignment = Alignment(horizontal='center')

    # Footer
    ws.merge_cells('B23:E23')
    ws['B23'] = "5746 E Apache St, Tulsa, OK 74115 | 918-835-6948 | sales@qaainc.com"
    ws['B23'].font = Font(size=9, color=DARK_GRAY)
    ws['B23'].alignment = Alignment(horizontal='center')

    return ws

# =============================================================================
# SUPPORTING SHEETS
# =============================================================================
def create_parts_master_sheet(wb):
    """Create Parts Master with 5H suffix entries"""
    ws = wb.create_sheet("Parts Master")

    headers = ['Part Number', 'Description', 'Cond', 'List Price', 'Core Charge', 'Stock', 'Lead Time', 'Notes']
    widths = [18, 35, 8, 12, 12, 8, 12, 30]

    for i, (header, width) in enumerate(zip(headers, widths)):
        col = get_column_letter(i + 1)
        ws[f'{col}1'] = header
        apply_style(ws[f'{col}1'], 'header')
        ws.column_dimensions[col].width = width

    # Parts data with CORRECT 5H suffix (NOT -5H!)
    parts = [
        ['10-382320', 'S-20/200 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-3823205H', 'S-20/200 Magneto - 500HR Inspection', 'Svc', 575, 0, 0, '3-5 Days', '500HR price'],
        ['10-382320-CO', 'S-20/200 Magneto - Cust Overhaul', 'CO', 895, 0, 0, '7-10 Days', 'Cust prop'],
        ['10-382320-OH', 'S-20/200 Magneto - OH Exchange', 'OH', 2495, 1000, 5, '5-7 Days', 'Exchange'],

        ['10-320680', 'S-1200 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-3206805H', 'S-1200 Magneto - 500HR Inspection', 'Svc', 495, 0, 0, '3-5 Days', '500HR price'],
        ['10-320680-CO', 'S-1200 Magneto - Cust Overhaul', 'CO', 750, 0, 0, '7-10 Days', 'Cust prop'],
        ['10-320680-OH', 'S-1200 Magneto - OH Exchange', 'OH', 2095, 800, 3, '5-7 Days', 'Exchange'],

        ['10-391088', 'D-2000/D-3000 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-3910885H', 'D-2000/D-3000 - 500HR Inspection', 'Svc', 695, 0, 0, '3-5 Days', '500HR price'],
        ['10-391088-CO', 'D-2000/D-3000 - Cust Overhaul', 'CO', 1150, 0, 0, '7-10 Days', 'Cust prop'],
        ['10-391088-OH', 'D-2000/D-3000 - OH Exchange', 'OH', 2850, 1100, 2, '5-7 Days', 'Exchange'],

        ['10-163020', 'Bendix S-20 Magneto', 'Base', 0, 0, 0, '', 'Base part'],
        ['10-1630205H', 'Bendix S-20 - 500HR Inspection', 'Svc', 625, 0, 0, '3-5 Days', '500HR price'],
        ['10-163020-CO', 'Bendix S-20 - Cust Overhaul', 'CO', 950, 0, 0, '7-10 Days', 'Cust prop'],
        ['10-163020-OH', 'Bendix S-20 - OH Exchange', 'OH', 2650, 950, 4, '5-7 Days', 'Exchange'],

        # Components
        ['LW-15473', 'Impulse Coupling Assembly', 'New', 425, 0, 12, 'In Stock', ''],
        ['10-357170', 'Coil Assembly', 'New', 185, 0, 20, 'In Stock', ''],
        ['10-357210', 'Capacitor', 'New', 45, 0, 50, 'In Stock', ''],
        ['10-51360', 'Distributor Block', 'New', 125, 0, 15, 'In Stock', ''],
        ['10-51350', 'Points Assembly', 'New', 85, 0, 30, 'In Stock', ''],
        ['10-51340', 'Bearing Set', 'New', 95, 0, 25, 'In Stock', ''],
        ['10-51330', 'Gear Assembly', 'New', 165, 0, 8, 'In Stock', ''],

        # Other products
        ['646942-CO', 'Governor - Cust Overhaul', 'CO', 1500, 0, 0, '14-21 Days', 'Vendor'],
        ['646942-OH', 'Governor - OH Exchange', 'OH', 3450, 1500, 2, '5-7 Days', ''],
        ['478615-CO', 'Turbocharger - Cust Overhaul', 'CO', 2800, 0, 0, '21-30 Days', 'Vendor'],
        ['478615-OH', 'Turbocharger - OH Exchange', 'OH', 5800, 2500, 1, '5-7 Days', ''],
    ]

    for row_num, row_data in enumerate(parts, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            apply_style(cell, 'data')
            if col_num in [4, 5]:
                cell.number_format = '$#,##0.00'

    return ws

def create_bom_sheet(wb):
    """Create BOM sheet"""
    ws = wb.create_sheet("BOM")

    headers = ['Parent P/N', 'Parent Description', 'Component P/N', 'Component Desc', 'Qty']
    for i, header in enumerate(headers):
        col = get_column_letter(i + 1)
        ws[f'{col}1'] = header
        apply_style(ws[f'{col}1'], 'header')
        ws.column_dimensions[col].width = 25

    bom_data = [
        ['10-382320', 'S-20/200 Magneto', '10-357170', 'Coil Assembly', 1],
        ['10-382320', 'S-20/200 Magneto', '10-357210', 'Capacitor', 1],
        ['10-382320', 'S-20/200 Magneto', '10-51360', 'Distributor Block', 1],
        ['10-382320', 'S-20/200 Magneto', 'LW-15473', 'Impulse Coupling', 1],
        ['10-320680', 'S-1200 Magneto', '10-357170', 'Coil Assembly', 1],
        ['10-320680', 'S-1200 Magneto', '10-357210', 'Capacitor', 1],
        ['10-391088', 'D-2000/D-3000', '10-357170', 'Coil Assembly', 1],
    ]

    for row_num, row_data in enumerate(bom_data, start=2):
        for col_num, value in enumerate(row_data, start=1):
            ws.cell(row=row_num, column=col_num, value=value)

    return ws

def create_config_sheet(wb):
    """Create Config sheet"""
    ws = wb.create_sheet("Config")

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35

    ws['A1'] = "SYSTEM CONFIGURATION"
    ws['A1'].font = Font(bold=True, size=14, color=NAVY)

    settings = [
        ('Quote Prefix', 'QAA', ''),
        ('Next Quote #', 1, 'Auto-increments'),
        ('Quote Valid Days', 30, ''),
        ('500HR Default Fee', 550, 'Fallback price'),
        ('500HR Findings Fee', 110, 'Per unit with findings'),
        ('Reminder Days', 7, 'Days before highlight'),
    ]

    for i, (name, value, note) in enumerate(settings, start=2):
        ws[f'A{i}'] = name
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
        apply_style(ws[f'B{i}'], 'input')
        ws[f'C{i}'] = note
        ws[f'C{i}'].font = Font(size=9, color=DARK_GRAY)

    # Tier discounts
    ws['A9'] = "TIER DISCOUNTS"
    ws['A9'].font = Font(bold=True, size=12, color=NAVY)

    tiers = [('Standard', 0), ('Tier 1', 0.07), ('Tier 2', 0.10), ('Tier 3', 0.13)]
    for i, (tier, discount) in enumerate(tiers, start=10):
        ws[f'A{i}'] = tier
        ws[f'B{i}'] = discount
        ws[f'B{i}'].number_format = '0%'

    return ws

# =============================================================================
# VBA MODULE GENERATOR
# =============================================================================
def create_vba_module_v2():
    """Generate updated VBA module with single Quote Tracker logging"""

    vba_code = '''Attribute VB_Name = "QAAQuoteSystem"
' ╔══════════════════════════════════════════════════════════════════════════════╗
' ║     QAA QUOTE SYSTEM v13.0 - PROFESSIONAL EDITION                            ║
' ║     Quality Aircraft Accessories - A Hartzell Engine Technologies Company    ║
' ║                                                                              ║
' ║     FEATURES:                                                                ║
' ║     - Uses 5H suffix (NOT -5H) for 500HR pricing                            ║
' ║     - Single Quote Tracker with follow-up workflow                          ║
' ║     - Reminder counting and status management                               ║
' ║     - Professional PDF generation                                           ║
' ╚══════════════════════════════════════════════════════════════════════════════╝

Option Explicit

' ═══════════════════════════════════════════════════════════════════════════════
' NAVIGATION
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub GoMenu()
    ThisWorkbook.Sheets("Menu").Activate
End Sub

Public Sub Go500HR()
    ThisWorkbook.Sheets("500HR").Activate
    ThisWorkbook.Sheets("500HR").Range("C6").Select
End Sub

Public Sub GoCustProp()
    ThisWorkbook.Sheets("Cust Property").Activate
    ThisWorkbook.Sheets("Cust Property").Range("C6").Select
End Sub

Public Sub GoExchange()
    ThisWorkbook.Sheets("Exchange").Activate
    ThisWorkbook.Sheets("Exchange").Range("B6").Select
End Sub

Public Sub GoTracker()
    ThisWorkbook.Sheets("Quote Tracker").Activate
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' PDF GENERATION - Professional Layout
' ═══════════════════════════════════════════════════════════════════════════════

Private Function GetPDFPath(quoteNum As String) As String
    Dim folder As String
    folder = ThisWorkbook.Path
    If folder = "" Then folder = Environ("USERPROFILE") & "\\Documents"
    GetPDFPath = folder & "\\QAA_Quote_" & Replace(Replace(quoteNum, "-", "_"), "/", "_") & "_" & Format(Date, "YYYYMMDD") & ".pdf"
End Function

Public Sub GenPDF500HR()
    Dim ws As Worksheet, pdfPath As String
    Set ws = ThisWorkbook.Sheets("500HR")

    If Trim(ws.Range("C6").Value) = "" Then
        MsgBox "Please enter customer email first.", vbExclamation, "QAA Quote System"
        Exit Sub
    End If

    pdfPath = GetPDFPath(ws.Range("C5").Value)

    Application.ScreenUpdating = False
    With ws.PageSetup
        .PrintArea = "$A$1:$J$37"
        .Orientation = xlPortrait
        .FitToPagesWide = 1
        .FitToPagesTall = 1
        .LeftMargin = Application.InchesToPoints(0.4)
        .RightMargin = Application.InchesToPoints(0.4)
        .TopMargin = Application.InchesToPoints(0.4)
        .BottomMargin = Application.InchesToPoints(0.4)
        .CenterHorizontally = True
        .PrintGridlines = False
    End With

    On Error GoTo PDFError
    ws.ExportAsFixedFormat xlTypePDF, pdfPath, xlQualityStandard
    Application.ScreenUpdating = True

    MsgBox "PDF Created:" & vbCrLf & vbCrLf & pdfPath, vbInformation, "QAA Quote System"

    If MsgBox("Send email with PDF attached?", vbYesNo + vbQuestion) = vbYes Then
        Email500HR pdfPath
    End If
    Exit Sub

PDFError:
    Application.ScreenUpdating = True
    MsgBox "PDF Error: " & Err.Description, vbCritical
End Sub

Public Sub GenPDFCustProp()
    Dim ws As Worksheet, pdfPath As String
    Set ws = ThisWorkbook.Sheets("Cust Property")

    If Trim(ws.Range("C6").Value) = "" Then
        MsgBox "Please enter customer email first.", vbExclamation
        Exit Sub
    End If

    pdfPath = GetPDFPath(ws.Range("C5").Value)

    Application.ScreenUpdating = False
    With ws.PageSetup
        .PrintArea = "$A$1:$I$36"
        .Orientation = xlPortrait
        .FitToPagesWide = 1
        .FitToPagesTall = 1
        .CenterHorizontally = True
    End With

    On Error GoTo PDFError
    ws.ExportAsFixedFormat xlTypePDF, pdfPath, xlQualityStandard
    Application.ScreenUpdating = True

    MsgBox "PDF Created:" & vbCrLf & vbCrLf & pdfPath, vbInformation

    If MsgBox("Send email with PDF attached?", vbYesNo + vbQuestion) = vbYes Then
        EmailCustProp pdfPath
    End If
    Exit Sub

PDFError:
    Application.ScreenUpdating = True
    MsgBox "PDF Error: " & Err.Description, vbCritical
End Sub

Public Sub GenPDFExchange()
    Dim ws As Worksheet, pdfPath As String
    Set ws = ThisWorkbook.Sheets("Exchange")

    If Trim(ws.Range("B6").Value) = "" Then
        MsgBox "Please enter customer email first.", vbExclamation
        Exit Sub
    End If

    pdfPath = GetPDFPath(ws.Range("B5").Value)

    Application.ScreenUpdating = False
    With ws.PageSetup
        .PrintArea = "$A$1:$J$31"
        .Orientation = xlLandscape
        .FitToPagesWide = 1
        .FitToPagesTall = 1
        .CenterHorizontally = True
    End With

    On Error GoTo PDFError
    ws.ExportAsFixedFormat xlTypePDF, pdfPath, xlQualityStandard
    Application.ScreenUpdating = True

    MsgBox "PDF Created:" & vbCrLf & vbCrLf & pdfPath, vbInformation

    If MsgBox("Send email with PDF attached?", vbYesNo + vbQuestion) = vbYes Then
        EmailExchange pdfPath
    End If
    Exit Sub

PDFError:
    Application.ScreenUpdating = True
    MsgBox "PDF Error: " & Err.Description, vbCritical
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' EMAIL GENERATION WITH QUOTE TRACKER LOGGING
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub Email500HR(Optional pdfPath As String = "")
    Dim OutApp As Object, OutMail As Object
    Dim ws As Worksheet
    Dim emailTo As String, subj As String, body As String
    Dim quoteNum As String, totalAmt As Double
    Dim needsApproval As String

    Set ws = ThisWorkbook.Sheets("500HR")
    emailTo = Trim(ws.Range("C6").Value)
    quoteNum = ws.Range("C5").Value
    totalAmt = ws.Range("H29").Value
    needsApproval = ws.Range("H6").Value

    If emailTo = "" Then
        MsgBox "Please enter customer email.", vbExclamation
        Exit Sub
    End If

    subj = "Quote " & quoteNum & " - 500HR Magneto Inspection - Quality Aircraft Accessories"

    body = "Dear Valued Customer," & vbCrLf & vbCrLf
    body = body & "Thank you for choosing Quality Aircraft Accessories for your magneto service needs." & vbCrLf & vbCrLf
    body = body & "Please find attached your quote for 500HR Magneto Inspection Service." & vbCrLf & vbCrLf
    body = body & String(50, "=") & vbCrLf
    body = body & "QUOTE SUMMARY" & vbCrLf
    body = body & String(50, "=") & vbCrLf
    body = body & "Quote #:     " & quoteNum & vbCrLf
    body = body & "Date:        " & Format(Date, "MMMM d, yyyy") & vbCrLf
    body = body & "Valid:       30 Days" & vbCrLf & vbCrLf
    body = body & "TOTAL:       " & Format(totalAmt, "$#,##0.00") & vbCrLf
    body = body & String(50, "=") & vbCrLf & vbCrLf

    If needsApproval = "Yes" Then
        body = body & "** APPROVAL REQUIRED **" & vbCrLf
        body = body & "Please reply with your approval to proceed." & vbCrLf & vbCrLf
    End If

    body = body & "Questions? Contact us anytime." & vbCrLf & vbCrLf
    body = body & "Best regards," & vbCrLf
    body = body & "Quality Aircraft Accessories" & vbCrLf
    body = body & "A Hartzell Engine Technologies Company" & vbCrLf
    body = body & "918-835-6948 | sales@qaainc.com | qaa.com"

    On Error GoTo EmailError
    Set OutApp = CreateObject("Outlook.Application")
    Set OutMail = OutApp.CreateItem(0)

    With OutMail
        .To = emailTo
        .Subject = subj
        .body = body
        If pdfPath <> "" Then
            If Dir(pdfPath) <> "" Then .Attachments.Add pdfPath
        End If
        .Display
    End With

    ' Log to Quote Tracker
    LogToTracker quoteNum, emailTo, "500HR Service", totalAmt, needsApproval, "No", "No", "Initial quote sent"

    MsgBox "Email created and logged to Quote Tracker!", vbInformation
    Exit Sub

EmailError:
    MsgBox "Email Error: " & Err.Description, vbCritical
End Sub

Public Sub EmailCustProp(Optional pdfPath As String = "")
    Dim OutApp As Object, OutMail As Object
    Dim ws As Worksheet
    Dim emailTo As String, subj As String, body As String
    Dim quoteNum As String, totalAmt As Double
    Dim needsAppr As String, needsShip As String, needsPay As String

    Set ws = ThisWorkbook.Sheets("Cust Property")
    emailTo = Trim(ws.Range("C6").Value)
    quoteNum = ws.Range("C5").Value
    totalAmt = ws.Range("G30").Value
    needsAppr = ws.Range("D7").Value
    needsShip = ws.Range("F7").Value
    needsPay = ws.Range("H7").Value

    If emailTo = "" Then
        MsgBox "Please enter customer email.", vbExclamation
        Exit Sub
    End If

    subj = "Quote " & quoteNum & " - Customer Property Overhaul - Quality Aircraft Accessories"

    body = "Dear Valued Customer," & vbCrLf & vbCrLf
    body = body & "Thank you for choosing Quality Aircraft Accessories." & vbCrLf & vbCrLf
    body = body & "Please find attached your quote for Customer Property Overhaul services." & vbCrLf & vbCrLf
    body = body & String(50, "=") & vbCrLf
    body = body & "QUOTE SUMMARY" & vbCrLf
    body = body & String(50, "=") & vbCrLf
    body = body & "Quote #:     " & quoteNum & vbCrLf
    body = body & "Date:        " & Format(Date, "MMMM d, yyyy") & vbCrLf
    body = body & "Valid:       30 Days" & vbCrLf & vbCrLf
    body = body & "TOTAL:       " & Format(totalAmt, "$#,##0.00") & vbCrLf
    body = body & String(50, "=") & vbCrLf & vbCrLf

    Dim req As String: req = ""
    If needsAppr = "Yes" Then req = req & "  - Approval to proceed" & vbCrLf
    If needsShip = "Yes" Then req = req & "  - Shipping information" & vbCrLf
    If needsPay = "Yes" Then req = req & "  - Payment method" & vbCrLf
    If req <> "" Then body = body & "TO PROCEED, WE NEED:" & vbCrLf & req & vbCrLf

    body = body & "Questions? Contact us anytime." & vbCrLf & vbCrLf
    body = body & "Best regards," & vbCrLf
    body = body & "Quality Aircraft Accessories" & vbCrLf
    body = body & "918-835-6948 | sales@qaainc.com | qaa.com"

    On Error GoTo EmailError
    Set OutApp = CreateObject("Outlook.Application")
    Set OutMail = OutApp.CreateItem(0)

    With OutMail
        .To = emailTo
        .Subject = subj
        .body = body
        If pdfPath <> "" Then
            If Dir(pdfPath) <> "" Then .Attachments.Add pdfPath
        End If
        .Display
    End With

    LogToTracker quoteNum, emailTo, "Cust Property", totalAmt, needsAppr, needsShip, needsPay, "Initial quote sent"

    MsgBox "Email created and logged to Quote Tracker!", vbInformation
    Exit Sub

EmailError:
    MsgBox "Email Error: " & Err.Description, vbCritical
End Sub

Public Sub EmailExchange(Optional pdfPath As String = "")
    Dim OutApp As Object, OutMail As Object
    Dim ws As Worksheet
    Dim emailTo As String, subj As String, body As String
    Dim quoteNum As String, totalAmt As Double
    Dim needsAppr As String, needsShip As String, needsPay As String

    Set ws = ThisWorkbook.Sheets("Exchange")
    emailTo = Trim(ws.Range("B6").Value)
    quoteNum = ws.Range("B5").Value
    totalAmt = ws.Range("F24").Value
    needsAppr = ws.Range("B7").Value
    needsShip = ws.Range("D7").Value
    needsPay = ws.Range("F7").Value

    If emailTo = "" Then
        MsgBox "Please enter customer email.", vbExclamation
        Exit Sub
    End If

    subj = "Quote " & quoteNum & " - Exchange/Sale - Quality Aircraft Accessories"

    body = "Dear Valued Customer," & vbCrLf & vbCrLf
    body = body & "Thank you for your interest in Quality Aircraft Accessories." & vbCrLf & vbCrLf
    body = body & "Please find attached your quote." & vbCrLf & vbCrLf
    body = body & String(50, "=") & vbCrLf
    body = body & "QUOTE SUMMARY" & vbCrLf
    body = body & String(50, "=") & vbCrLf
    body = body & "Quote #:     " & quoteNum & vbCrLf
    body = body & "Date:        " & Format(Date, "MMMM d, yyyy") & vbCrLf
    body = body & "Valid:       30 Days" & vbCrLf & vbCrLf
    body = body & "TOTAL:       " & Format(totalAmt, "$#,##0.00") & vbCrLf
    body = body & String(50, "=") & vbCrLf & vbCrLf

    Dim req As String: req = ""
    If needsAppr = "Yes" Then req = req & "  - Approval" & vbCrLf
    If needsShip = "Yes" Then req = req & "  - Shipping info" & vbCrLf
    If needsPay = "Yes" Then req = req & "  - Payment method" & vbCrLf
    If req <> "" Then body = body & "TO PROCEED:" & vbCrLf & req & vbCrLf

    body = body & "Questions? Contact us anytime." & vbCrLf & vbCrLf
    body = body & "Best regards," & vbCrLf
    body = body & "Quality Aircraft Accessories" & vbCrLf
    body = body & "918-835-6948 | sales@qaainc.com | qaa.com"

    On Error GoTo EmailError
    Set OutApp = CreateObject("Outlook.Application")
    Set OutMail = OutApp.CreateItem(0)

    With OutMail
        .To = emailTo
        .Subject = subj
        .body = body
        If pdfPath <> "" Then
            If Dir(pdfPath) <> "" Then .Attachments.Add pdfPath
        End If
        .Display
    End With

    LogToTracker quoteNum, emailTo, "Exchange/Sale", totalAmt, needsAppr, needsShip, needsPay, "Initial quote sent"

    MsgBox "Email created and logged to Quote Tracker!", vbInformation
    Exit Sub

EmailError:
    MsgBox "Email Error: " & Err.Description, vbCritical
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' QUOTE TRACKER LOGGING
' ═══════════════════════════════════════════════════════════════════════════════

Private Sub LogToTracker(quoteNum As String, emailTo As String, qType As String, _
                         total As Double, needsAppr As String, needsShip As String, _
                         needsPay As String, notes As String)

    Dim wsTracker As Worksheet, wsConfig As Worksheet
    Dim nr As Long
    Dim existingRow As Long

    On Error GoTo LogError

    Set wsTracker = ThisWorkbook.Sheets("Quote Tracker")
    Set wsConfig = ThisWorkbook.Sheets("Config")

    ' Check if quote already exists (for follow-ups)
    existingRow = FindQuoteRow(quoteNum)

    If existingRow > 0 Then
        ' Update existing entry (increment reminder count)
        wsTracker.Cells(existingRow, 9).Value = wsTracker.Cells(existingRow, 9).Value + 1  ' Reminders
        wsTracker.Cells(existingRow, 10).Value = Date  ' Last Contact
        wsTracker.Cells(existingRow, 13).Value = wsTracker.Cells(existingRow, 13).Value & vbCrLf & _
            Format(Now, "MM/DD HH:MM") & " - " & notes
    Else
        ' New entry
        nr = wsTracker.Cells(wsTracker.Rows.Count, 1).End(xlUp).Row + 1
        If nr < 4 Then nr = 4  ' Start after headers

        wsTracker.Cells(nr, 1).Value = quoteNum                     ' Quote #
        wsTracker.Cells(nr, 2).Value = Date                          ' Created
        wsTracker.Cells(nr, 3).Value = emailTo                       ' Customer Email
        wsTracker.Cells(nr, 4).Value = qType                         ' Type
        wsTracker.Cells(nr, 5).Value = total                         ' Amount
        wsTracker.Cells(nr, 6).Value = IIf(needsAppr = "Yes", "Yes", "No")  ' Approval
        wsTracker.Cells(nr, 7).Value = IIf(needsShip = "Yes", "Yes", "No")  ' Shipping
        wsTracker.Cells(nr, 8).Value = IIf(needsPay = "Yes", "Yes", "No")   ' Payment
        wsTracker.Cells(nr, 9).Value = 1                             ' Reminders (first contact)
        wsTracker.Cells(nr, 10).Value = Date                         ' Last Contact
        wsTracker.Cells(nr, 12).Value = "OPEN"                       ' Status
        wsTracker.Cells(nr, 13).Value = Format(Now, "MM/DD HH:MM") & " - " & notes  ' Notes

        ' Increment quote number
        wsConfig.Range("B3").Value = wsConfig.Range("B3").Value + 1
    End If

    Exit Sub

LogError:
    MsgBox "Tracker Log Error: " & Err.Description, vbExclamation
End Sub

Private Function FindQuoteRow(quoteNum As String) As Long
    Dim wsTracker As Worksheet
    Dim r As Long

    Set wsTracker = ThisWorkbook.Sheets("Quote Tracker")
    FindQuoteRow = 0

    For r = 4 To wsTracker.Cells(wsTracker.Rows.Count, 1).End(xlUp).Row
        If wsTracker.Cells(r, 1).Value = quoteNum Then
            FindQuoteRow = r
            Exit Function
        End If
    Next r
End Function

' ═══════════════════════════════════════════════════════════════════════════════
' REMINDER & STATUS MANAGEMENT
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub SendReminder()
    ' Send a reminder for the currently selected quote in tracker
    Dim wsTracker As Worksheet
    Dim r As Long
    Dim quoteNum As String, emailTo As String

    Set wsTracker = ThisWorkbook.Sheets("Quote Tracker")
    r = ActiveCell.Row

    If r < 4 Then
        MsgBox "Please select a quote row first.", vbExclamation
        Exit Sub
    End If

    quoteNum = wsTracker.Cells(r, 1).Value
    emailTo = wsTracker.Cells(r, 3).Value

    If quoteNum = "" Then
        MsgBox "No quote selected.", vbExclamation
        Exit Sub
    End If

    ' Update tracker
    wsTracker.Cells(r, 9).Value = wsTracker.Cells(r, 9).Value + 1  ' Increment reminders
    wsTracker.Cells(r, 10).Value = Date  ' Update last contact
    wsTracker.Cells(r, 13).Value = wsTracker.Cells(r, 13).Value & vbCrLf & _
        Format(Now, "MM/DD HH:MM") & " - Reminder sent"

    MsgBox "Reminder logged for " & quoteNum & vbCrLf & _
           "Total reminders: " & wsTracker.Cells(r, 9).Value, vbInformation
End Sub

Public Sub CloseQuote()
    ' Mark the selected quote as CLOSED
    Dim wsTracker As Worksheet
    Dim r As Long

    Set wsTracker = ThisWorkbook.Sheets("Quote Tracker")
    r = ActiveCell.Row

    If r < 4 Then
        MsgBox "Please select a quote row first.", vbExclamation
        Exit Sub
    End If

    wsTracker.Cells(r, 12).Value = "CLOSED"
    wsTracker.Cells(r, 13).Value = wsTracker.Cells(r, 13).Value & vbCrLf & _
        Format(Now, "MM/DD HH:MM") & " - Quote CLOSED"

    MsgBox "Quote marked as CLOSED.", vbInformation
End Sub

' ═══════════════════════════════════════════════════════════════════════════════
' CLEAR FORMS
' ═══════════════════════════════════════════════════════════════════════════════

Public Sub Clear500HR()
    Dim ws As Worksheet, r As Integer, c As Integer
    Set ws = ThisWorkbook.Sheets("500HR")

    If MsgBox("Clear 500HR form?", vbYesNo + vbQuestion) = vbNo Then Exit Sub

    ws.Range("C6").ClearContents
    ws.Range("H6").Value = "No"

    For r = 9 To 14
        ws.Cells(r, 2).ClearContents
        ws.Cells(r, 3).ClearContents
        ws.Cells(r, 6).ClearContents
    Next r

    For r = 18 To 23
        For c = 2 To 9
            ws.Cells(r, c).Value = ChrW(9744)
        Next c
    Next r

    ws.Range("B32").ClearContents
    MsgBox "Cleared!", vbInformation
End Sub

Public Sub ClearCustProp()
    Dim ws As Worksheet, r As Integer
    Set ws = ThisWorkbook.Sheets("Cust Property")

    If MsgBox("Clear Cust Property form?", vbYesNo + vbQuestion) = vbNo Then Exit Sub

    ws.Range("C6").ClearContents
    ws.Range("H5").ClearContents
    ws.Range("G6").Value = "Standard"
    ws.Range("D7").Value = "No"
    ws.Range("F7").Value = "No"
    ws.Range("H7").Value = "No"

    For r = 10 To 13
        ws.Cells(r, 2).ClearContents
        ws.Cells(r, 5).ClearContents
    Next r

    For r = 17 To 26
        ws.Cells(r, 1).ClearContents
        ws.Cells(r, 2).ClearContents
        ws.Cells(r, 4).ClearContents
        ws.Cells(r, 6).ClearContents
        ws.Cells(r, 8).ClearContents
    Next r

    ws.Range("B33").ClearContents
    MsgBox "Cleared!", vbInformation
End Sub

Public Sub ClearExchange()
    Dim ws As Worksheet, r As Integer
    Set ws = ThisWorkbook.Sheets("Exchange")

    If MsgBox("Clear Exchange form?", vbYesNo + vbQuestion) = vbNo Then Exit Sub

    ws.Range("B6").ClearContents
    ws.Range("G6").Value = "Standard"
    ws.Range("B7").Value = "No"
    ws.Range("D7").Value = "No"
    ws.Range("F7").Value = "No"
    ws.Range("H7").Value = "No"

    For r = 10 To 21
        ws.Cells(r, 1).ClearContents
        ws.Cells(r, 4).ClearContents
    Next r

    ws.Range("A27").ClearContents
    MsgBox "Cleared!", vbInformation
End Sub

Public Sub UpdateDate()
    ThisWorkbook.Sheets("Cust Property").Range("H5").Value = Date
    MsgBox "Date updated!", vbInformation
End Sub
'''
    return vba_code

# =============================================================================
# MAIN BUILD FUNCTION
# =============================================================================
def build_workbook_v2(output_dir, logo_path):
    """Build the complete QAA Quote System v13.0"""

    print("=" * 60)
    print("QAA QUOTE SYSTEM v13.0 - PROFESSIONAL EDITION")
    print("=" * 60)

    # Create buttons
    print("\n[1/9] Creating professional button images...")
    create_professional_buttons(output_dir)

    # Create workbook
    print("\n[2/9] Creating workbook...")
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    print("[3/9] Creating Menu...")
    create_menu_sheet(wb, logo_path)

    print("[4/9] Creating 500HR sheet (5H suffix formulas)...")
    create_500hr_sheet(wb)

    print("[5/9] Creating Cust Property sheet...")
    create_cust_property_sheet(wb)

    print("[6/9] Creating Exchange sheet...")
    create_exchange_sheet(wb)

    print("[7/9] Creating Quote Tracker (single intelligent log)...")
    create_quote_tracker_sheet(wb)

    print("[8/9] Creating supporting sheets...")
    create_parts_master_sheet(wb)
    create_bom_sheet(wb)
    create_config_sheet(wb)

    # Add buttons and logo
    print("[9/9] Adding images...")

    # Add logo to sheets
    if logo_path and os.path.exists(logo_path):
        for sheet in ['500HR', 'Cust Property', 'Exchange']:
            try:
                logo = XLImage(logo_path)
                logo.width = 140
                logo.height = 50
                wb[sheet].add_image(logo, 'H1')
            except:
                pass

    # Save
    xlsx_path = os.path.join(output_dir, 'QAA_Quote_System_v13.xlsx')
    wb.save(xlsx_path)
    print(f"\n  Saved: {xlsx_path}")

    # VBA module
    vba_path = os.path.join(output_dir, 'QAA_VBA_MODULE_v13.bas')
    with open(vba_path, 'w') as f:
        f.write(create_vba_module_v2())
    print(f"  VBA Module: {vba_path}")

    print("\n" + "=" * 60)
    print("BUILD COMPLETE!")
    print("=" * 60)
    print("\nKEY IMPROVEMENTS:")
    print("  - Professional PDF layout with proper spacing")
    print("  - Single Quote Tracker with follow-up workflow")
    print("  - RED highlighting for quotes needing follow-up (7+ days)")
    print("  - Reminder count tracking")
    print("  - Status management (OPEN/CLOSED)")
    print("  - 5H suffix formulas (NOT -5H)")

    return xlsx_path, vba_path

if __name__ == "__main__":
    output_dir = "/home/user/Test"
    logo_path = "/home/user/Test/qaa_logo_header.png"

    if not os.path.exists(logo_path):
        logo_path = "/home/user/Test/extracted/package/qaa_logo.png"

    build_workbook_v2(output_dir, logo_path)
